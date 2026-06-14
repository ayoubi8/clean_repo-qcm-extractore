import asyncio
import builtins
import json
import sys
import os
import time
from datetime import datetime
import yaml
from pathlib import Path
from dotenv import load_dotenv
from fastapi import FastAPI, WebSocket, BackgroundTasks, HTTPException, UploadFile, File, Depends, Request, Header

from fastapi.responses import FileResponse, RedirectResponse
import shutil
import mimetypes
import pickle
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleAuthRequest
from googleapiclient.discovery import build
from fastapi.middleware.cors import CORSMiddleware

load_dotenv("/app/.env" if Path("/app").exists() else ".env", override=False)
load_dotenv(override=False)


# Setup path to include the root /app (so modules can be imported)
sys.path.insert(0, "/app")

from job_manager import job_manager
from log_capture import LogCapture
from project_manager import get_or_create, list_projects, step_output_exists, get_weekly_costs
from env_manager import read_env, mask, write_env_keys, EDITABLE_KEYS
from auth import (get_current_user, require_admin, load_users,
                     find_user_by_email, hash_password, verify_password, rehash_if_legacy,
                     create_access_token, add_user, update_user_field, delete_user,
                     check_rate_limit,
                     create_refresh_token, verify_and_rotate_refresh_token, revoke_all_refresh_tokens,
                     ADMIN_EMAIL, ADMIN_PASSWORD, ensure_admin_exists)
import uuid
from supabase_client import get_supabase
from storage_client import (
    write_file, write_bytes_file, read_file, read_bytes_file,
    file_exists, list_files, delete_prefix, get_signed_url,
    write_pickle, read_pickle
)

app = FastAPI(title="QCM Extractor API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
async def _startup():
    """Seed admin user and decode Google credentials on every container start."""
    ensure_admin_exists()

    # Restore .env from Supabase Storage for persistence across restarts
    try:
        if file_exists("config/.env"):
            env_content = read_file("config/.env")
            dest_path = Path("/app/.env") if Path("/app").exists() else Path(__file__).parent / ".env"
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            dest_path.write_text(env_content)
            print("[STARTUP] ✅ Restored .env from Supabase Storage")
            load_dotenv(str(dest_path), override=True)
            print("[STARTUP] ✅ Reloaded env variables with override=True")
    except Exception as e:
        print(f"[STARTUP] ❌ Failed to restore .env from Supabase Storage: {e}")


    # Decode Google OAuth client secret from base64 env var → write to disk
    google_secret_b64 = os.environ.get("GOOGLE_CLIENT_SECRET_B64", "")
    if google_secret_b64:
        import base64
        try:
            secret_bytes = base64.b64decode(google_secret_b64)
            secret_path = Path(GOOGLE_CLIENT_SECRET_PATH)
            secret_path.parent.mkdir(parents=True, exist_ok=True)
            secret_path.write_bytes(secret_bytes)
            print(f"[STARTUP] ✅ Google client secret written to {GOOGLE_CLIENT_SECRET_PATH}")
        except Exception as e:
            print(f"[STARTUP] ❌ Could not decode GOOGLE_CLIENT_SECRET_B64: {e}")
    else:
        print("[STARTUP] ⚠️ GOOGLE_CLIENT_SECRET_B64 not set — Google Sheets integration will NOT work")

    # Log Google Sheets readiness
    print(f"[STARTUP] Google OAuth redirect URI: {GOOGLE_REDIRECT_URI}")
    if Path(GOOGLE_CLIENT_SECRET_PATH).exists():
        print(f"[STARTUP] ✅ Google client secret file exists at {GOOGLE_CLIENT_SECRET_PATH}")
    else:
        print(f"[STARTUP] ❌ Google client secret file MISSING at {GOOGLE_CLIENT_SECRET_PATH}")


def _migrate_legacy_projects():
    """Move folders from /app/output/ to /app/output/admin_email/ if they aren't isolated yet."""
    base_dir = Path("/app/output")
    if not base_dir.exists():
        return

    admin_isolated_dir = base_dir / ADMIN_EMAIL
    migrated_count = 0

    try:
        for item in base_dir.iterdir():
            if not item.is_dir():
                continue

            name = item.name
            # Skip already isolated folders (contain @), special folders, and hidden folders
            if "@" in name or name == "_history" or name.startswith("."):
                continue

            # This is a legacy project folder. Move it to the admin's account.
            admin_isolated_dir.mkdir(parents=True, exist_ok=True)
            dest = admin_isolated_dir / name

            if dest.exists():
                print(f"[MIGRATION] Target '{dest}' already exists, skipping '{name}'")
                continue

            shutil.move(str(item), str(dest))
            print(f"[MIGRATION] Moved legacy project '{name}' to admin account ({ADMIN_EMAIL})")
            migrated_count += 1

        if migrated_count == 0:
            print("[MIGRATION] No legacy projects found, skipping.")
        else:
            print(f"[MIGRATION] Migration complete. {migrated_count} project(s) moved.")
    except Exception as e:
        print(f"[MIGRATION] Error during migration: {e}")

    # --- Phase 2: email-folder → UUID-folder migration ---
    try:
        users = load_users()
        for u in users:
            src = base_dir / u["email"]
            dst = base_dir / u["id"]
            if src.exists() and not dst.exists():
                src.rename(dst)
                print(f"[UUID-MIGRATION] {u['email']} → {u['id']}")
        # Admin folder
        admin_src = base_dir / ADMIN_EMAIL
        admin_dst = base_dir / "admin"
        if admin_src.exists() and not admin_dst.exists():
            admin_src.rename(admin_dst)
            print("[UUID-MIGRATION] admin email folder → admin")
    except Exception as e:
        print(f"[UUID-MIGRATION] Error: {e}")

@app.on_event("startup")
async def startup_event():
    _migrate_legacy_projects()

# --- Auth Endpoints ---

@app.post("/auth/register")
def register(body: dict, request: Request):
    check_rate_limit(request.client.host, "register")
    email = body.get("email")
    password = body.get("password")
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password required")

    if email == ADMIN_EMAIL or find_user_by_email(email):
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hash_password(password),
        "is_approved": False,
        "is_admin": False,
        "api_key": "",
        "allowed_models": {},
        "created_at": datetime.utcnow().isoformat()
    }
    add_user(new_user)
    return {"message": "Registration successful. Awaiting admin approval."}

@app.post("/auth/login")
def login(body: dict, request: Request):
    check_rate_limit(request.client.host, "login")
    email    = body.get("email", "").strip().lower()
    password = body.get("password", "")

    user = find_user_by_email(email)
    if not user or not verify_password(password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Transparently upgrade legacy hash to bcrypt
    new_hash = rehash_if_legacy(password, user["password_hash"])
    if new_hash:
        update_user_field(user["id"], "password_hash", new_hash)

    if not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="PENDING_APPROVAL")

    is_admin      = user.get("is_admin", False) or (email == ADMIN_EMAIL)
    token         = create_access_token(user["id"], user["email"], is_admin)
    refresh_token = create_refresh_token(user["id"])
    return {
        "access_token":  token,
        "refresh_token": refresh_token,
        "token_type":    "bearer",
        "is_admin":      is_admin,
        "email":         user["email"]
    }

@app.get("/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    # Return user info without password hash
    return {k: v for k, v in user.items() if k != "password_hash"}

@app.post("/auth/refresh")
def refresh_access(body: dict):
    raw_token = body.get("refresh_token")
    if not raw_token:
        raise HTTPException(status_code=400, detail="refresh_token required")
    new_access, new_refresh = verify_and_rotate_refresh_token(raw_token)
    return {"access_token": new_access, "refresh_token": new_refresh, "token_type": "bearer"}

@app.post("/auth/logout")
def logout_user(user: dict = Depends(get_current_user)):
    revoke_all_refresh_tokens(user["id"])
    return {"ok": True}

# --- Admin Endpoints ---

@app.get("/admin/users")
def list_admin_users(admin: dict = Depends(require_admin)):
    users = load_users()
    enriched_users = []

    # Inject Admin synthetic record — folder is /app/output/admin
    admin_dir = Path("/app/output/admin")
    admin_project_count = 0
    admin_total_cost = 0.0
    admin_total_tokens = 0

    if admin_dir.exists():
        admin_projects = [d for d in admin_dir.iterdir() if d.is_dir() and not d.name.startswith(("_", "."))]
        admin_project_count = len(admin_projects)
        for p_dir in admin_projects:
            cost_file = p_dir / "total_costs.json"
            if cost_file.exists():
                try:
                    data = json.loads(cost_file.read_text())
                    summary = data.get("summary", data)
                    admin_total_cost += summary.get("total_cost", 0.0)
                    admin_total_tokens += summary.get("total_tokens", 0)
                except: pass

    admin_record = {
        "id": "admin",
        "email": ADMIN_EMAIL,
        "is_approved": True,
        "is_admin": True,
        "api_key": "",
        "allowed_models": {},
        "created_at": "2024-01-01T00:00:00",
        "project_count": admin_project_count,
        "total_cost": round(admin_total_cost, 4),
        "total_tokens": admin_total_tokens,
    }
    enriched_users.append(admin_record)

    for user in users:
        user_dir = Path(f"/app/output/{user['id']}")

        project_count = 0
        total_cost = 0.0
        total_tokens = 0

        if user_dir.exists():
            projects = [d for d in user_dir.iterdir() if d.is_dir() and not d.name.startswith("_") and not d.name.startswith(".")]
            project_count = len(projects)

            for p_dir in projects:
                cost_file = p_dir / "total_costs.json"
                if cost_file.exists():
                    try:
                        data = json.loads(cost_file.read_text())
                        summary = data.get("summary", data)
                        total_cost += summary.get("total_cost", 0.0)
                        total_tokens += summary.get("total_tokens", 0)
                    except: pass

        user_copy = {k: v for k, v in user.items() if k != "password_hash"}
        user_copy.update({
            "project_count": project_count,
            "total_cost": round(total_cost, 4),
            "total_tokens": total_tokens
        })
        enriched_users.append(user_copy)

    return enriched_users

@app.patch("/admin/users/{uid}/approve")
def approve_user(uid: str, admin: dict = Depends(require_admin)):
    if not find_user_by_id(uid):
        raise HTTPException(status_code=404, detail="User not found")
    update_user_field(uid, "is_approved", True)
    return {"ok": True}

@app.patch("/admin/users/{uid}/reject")
def reject_user(uid: str, admin: dict = Depends(require_admin)):
    if not find_user_by_id(uid):
        raise HTTPException(status_code=404, detail="User not found")
    delete_user(uid)
    return {"ok": True}

@app.patch("/admin/users/{uid}/api-key")
def set_user_api_key(uid: str, body: dict, admin: dict = Depends(require_admin)):
    api_key = body.get("api_key", "")
    if not find_user_by_id(uid):
        raise HTTPException(status_code=404, detail="User not found")
    update_user_field(uid, "api_key", api_key)
    return {"ok": True}

@app.patch("/admin/users/{uid}/models")
def set_user_models(uid: str, body: dict, admin: dict = Depends(require_admin)):
    allowed_models = body.get("allowed_models", {})
    if not find_user_by_id(uid):
        raise HTTPException(status_code=404, detail="User not found")
    update_user_field(uid, "allowed_models", allowed_models)
    return {"ok": True}

@app.get("/admin/stats")
def get_admin_stats(admin: dict = Depends(require_admin)):
    users_data = list_admin_users(admin)
    
    total_users = len(users_data)
    total_projects = sum(u["project_count"] for u in users_data)
    total_cost = sum(u["total_cost"] for u in users_data)
    total_tokens = sum(u["total_tokens"] for u in users_data)
    
    return {
        "total_users": total_users,
        "total_projects": total_projects,
        "total_cost": round(total_cost, 4),
        "total_tokens": total_tokens,
        "per_user": users_data
    }

@app.get("/admin/users/{uid}/projects")
def get_user_projects(uid: str, admin: dict = Depends(require_admin)):
    # uid is either "admin" or a UUID — use directly as folder name
    if uid != "admin":
        from auth import find_user_by_id
        if not find_user_by_id(uid):
            raise HTTPException(status_code=404, detail="User not found")

    user_dir = Path(f"/app/output/{uid}")
    if not user_dir.exists():
        return {"projects": []}
    
    projects = []
    for d in user_dir.iterdir():
        if d.is_dir() and not d.name.startswith(("_", ".")):
            total_cost = 0.0
            total_tokens = 0
            cost_file = d / "total_costs.json"
            if cost_file.exists():
                try:
                    data = json.loads(cost_file.read_text())
                    summary = data.get("summary", data)
                    total_cost = summary.get("total_cost", 0.0)
                    total_tokens = summary.get("total_tokens", 0)
                except: pass
            
            last_modified = datetime.utcfromtimestamp(
                os.path.getmtime(str(d))
            ).isoformat()
            
            projects.append({
                "name": d.name,
                "last_modified": last_modified,
                "total_cost": round(total_cost, 4),
                "total_tokens": total_tokens,
            })
    
    return {"projects": sorted(projects, key=lambda p: p["last_modified"], reverse=True)}

def _apply_user_env(user: dict):
    key = user.get("api_key", "")
    if key:
        os.environ["GEMINI_API_KEY"] = key
        os.environ["OPENAI_API_KEY"] = key
        os.environ["ANTHROPIC_API_KEY"] = key
        os.environ["OPENROUTER_API_KEY"] = key


# --- Project Endpoints ---

@app.get("/projects")
def get_projects(user: dict = Depends(get_current_user)):
    return {"projects": list_projects(user["id"])}

@app.post("/projects")
def create_project(body: dict, user: dict = Depends(get_current_user)):
    name = body.get("name")
    if not name:
        raise HTTPException(status_code=400, detail="Project name required")
    get_or_create(name, user["id"])

    pdf_path = body.get("pdf_path", "")
    project_meta = json.dumps({"name": name, "pdf_path": pdf_path})

    # Write project.json locally (pipeline modules need it)
    project_dir = Path(f"/app/output/{user['id']}/{name}")
    project_dir.mkdir(parents=True, exist_ok=True)
    (project_dir / "project.json").write_text(project_meta)

    # Also persist to Supabase Storage (survives container restarts)
    try:
        write_file(f"{user['id']}/{name}/project.json", project_meta)
    except Exception as e:
        print(f"[STORAGE] project.json upload failed: {e}")

    return {
        "name": name,
        "pdf_path": pdf_path,
        "last_step": 0,
        "last_modified": "",
        "total_tokens": 0
    }

# --- Google Sheets Config ---
GOOGLE_CLIENT_SECRET_PATH = "/app/google_client_secret.json" if Path("/app").exists() else str(Path(__file__).parent / "google_client_secret.json")
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive",
]
GOOGLE_REDIRECT_URI = os.environ.get(
    "GOOGLE_REDIRECT_URI",
    "https://ayoubi8-qcm-extractor.hf.space/oauth2callback"   # production default
)


def _get_google_creds(user_id: str):
    """Load saved token from Supabase Storage or local FS. Returns None if not authorized."""
    storage_path = f"{user_id}/.google_token.pickle"

    # Try Supabase Storage first
    try:
        creds = read_pickle(storage_path)
        if creds and creds.valid:
            print(f"[GOOGLE] Valid creds loaded from Supabase for {user_id}")
            return creds
        if creds and creds.expired and creds.refresh_token:
            print(f"[GOOGLE] Creds expired for {user_id}, refreshing...")
            creds.refresh(GoogleAuthRequest())
            try:
                write_pickle(storage_path, creds)
            except Exception:
                pass
            print(f"[GOOGLE] Creds refreshed successfully for {user_id}")
            return creds
        print(f"[GOOGLE] Creds from Supabase exist but not valid/refreshable for {user_id}")
    except Exception as e:
        print(f"[GOOGLE] Supabase token load failed for {user_id}: {e}")

    # Fallback to local filesystem
    token_path = Path(f"/app/output/{user_id}/.google_token.pickle")
    if token_path.exists():
        with open(token_path, "rb") as f:
            try:
                creds = pickle.load(f)
                if creds and creds.valid:
                    print(f"[GOOGLE] Valid creds loaded from local FS for {user_id}")
                    return creds
                if creds and creds.expired and creds.refresh_token:
                    print(f"[GOOGLE] Local creds expired for {user_id}, refreshing...")
                    creds.refresh(GoogleAuthRequest())
                    with open(token_path, "wb") as f2:
                        pickle.dump(creds, f2)
                    print(f"[GOOGLE] Local creds refreshed successfully for {user_id}")
                    return creds
            except Exception as e:
                print(f"[GOOGLE] Error loading local google token for {user_id}: {e}")
    else:
        print(f"[GOOGLE] No token found (Supabase or local) for {user_id}")
    return None

# --- PDF Upload & View Endpoints ---

@app.get("/projects/{name}/pdf")
def serve_project_pdf(name: str, user: dict = Depends(get_current_user)):
    """Serve the project's source PDF — from Supabase Storage (signed URL) or local FS fallback."""
    storage_path = f"{user['id']}/{name}/source.pdf"

    # Try Supabase Storage first
    try:
        if file_exists(storage_path):
            signed_url = get_signed_url(storage_path)
            return RedirectResponse(signed_url)
    except Exception as e:
        print(f"[STORAGE] signed URL failed: {e}")

    # Fallback to local filesystem
    pdf_path = Path(f"/app/output/{user['id']}/{name}/source.pdf")
    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail="PDF not found")
    return FileResponse(str(pdf_path), media_type="application/pdf")

@app.get("/auth/google")
def start_google_auth(
    project: str = "", step: str = "", filename: str = "",
    token: str = "",          # JWT passed as query param from browser redirects
    authorization: str = Header(None),
):
    """Redirect user to Google consent screen. Accepts JWT via header OR ?token= query param."""
    from auth import decode_token, find_user_by_id
    # Resolve JWT — prefer header, fall back to query param
    raw_token = None
    if authorization and authorization.startswith("Bearer "):
        raw_token = authorization.split(" ")[1]
    elif token:
        raw_token = token
    if not raw_token:
        raise HTTPException(status_code=401, detail="Missing or invalid authorization header")
    payload = decode_token(raw_token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user_id = payload.get("sub", "admin")

    flow = Flow.from_client_secrets_file(
        GOOGLE_CLIENT_SECRET_PATH,
        scopes=GOOGLE_SCOPES,
        redirect_uri=GOOGLE_REDIRECT_URI
    )
    auth_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=f"{project}||{step}||{filename}||{user_id}"
    )
    return RedirectResponse(auth_url)

@app.get("/oauth2callback")
def google_oauth_callback(code: str, state: str = ""):
    """Handle Google OAuth callback, save token, redirect back to frontend."""
    print(f"[OAUTH] Callback received. redirect_uri={GOOGLE_REDIRECT_URI}")
    try:
        os.environ["OAUTHLIB_RELAX_TOKEN_SCOPE"] = "1"  # Allow Google to grant extra scopes
        flow = Flow.from_client_secrets_file(
            GOOGLE_CLIENT_SECRET_PATH,
            scopes=GOOGLE_SCOPES,
            redirect_uri=GOOGLE_REDIRECT_URI
        )
        flow.fetch_token(code=code)
        creds = flow.credentials
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Google token exchange failed: {str(e)}")

    # Parse state: project||step||filename||user_id
    parts = state.split("||")
    project  = parts[0] if len(parts) > 0 else ""
    step     = parts[1] if len(parts) > 1 else ""
    filename = parts[2] if len(parts) > 2 else ""
    user_id  = parts[3] if len(parts) > 3 else "admin"

    # Save to Supabase Storage (primary — survives restarts)
    try:
        write_pickle(f"{user_id}/.google_token.pickle", creds)
        print(f"[OAUTH] Token saved to Supabase for user {user_id}")
    except Exception as e:
        print(f"[STORAGE] Google token Supabase upload failed: {e}")

    # Also save to local filesystem (fallback for pipeline compatibility)
    try:
        token_path = Path(f"/app/output/{user_id}/.google_token.pickle")
        token_path.parent.mkdir(parents=True, exist_ok=True)
        with open(token_path, "wb") as f:
            pickle.dump(creds, f)
    except Exception as e:
        print(f"[OAUTH] Local token save failed: {e}")

    # Token saved — redirect back to frontend to retry the upload via /open-sheets
    frontend_url = os.environ.get("FRONTEND_URL", "https://qcm-extractor-frontend.vercel.app")
    from urllib.parse import quote as _url_quote
    return RedirectResponse(
        f"{frontend_url}/pipeline"
        f"?sheets_pending=1"
        f"&project={_url_quote(project, safe='')}"
        f"&step={_url_quote(step, safe='')}"
        f"&filename={_url_quote(filename, safe='')}"
    )







@app.post("/projects/{name}/pdf")
async def upload_project_pdf(name: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a PDF — dual-write to Supabase Storage and local FS for pipeline compatibility."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    content = await file.read()
    storage_pdf_path = f"{user['id']}/{name}/source.pdf"

    # Upload to Supabase Storage (cloud-persistent)
    # Store the absolute local path in project.json so that after a container
    # restart, list_projects restores the correct path and Step 1 can find the file.
    project_dir = Path(f"/app/output/{user['id']}/{name}")
    project_dir.mkdir(parents=True, exist_ok=True)
    local_pdf_path = project_dir / "source.pdf"
    local_pdf_path.write_bytes(content)
    internal_path = str(local_pdf_path)          # /app/output/{uid}/{name}/source.pdf

    try:
        write_bytes_file(storage_pdf_path, content)
        # ✅ Store the LOCAL absolute path so restored project.json is correct
        write_file(
            f"{user['id']}/{name}/project.json",
            json.dumps({"name": name, "pdf_path": internal_path})
        )
    except Exception as e:
        print(f"[STORAGE] PDF upload to Supabase failed: {e}")

    # Write the same project.json locally
    (project_dir / "project.json").write_text(
        json.dumps({"name": name, "pdf_path": internal_path})
    )

    return {"pdf_path": internal_path, "size_bytes": len(content)}


# --- Reference Database Management Endpoints ---

def _get_user_db_id(user: dict) -> str:
    """Helper to return the actual database UUID for the user (resolves 'admin' placeholder to UUID)."""
    from auth import find_user_by_email, ADMIN_EMAIL
    user_id = user.get("id", "")
    if user_id == "admin":
        admin_record = find_user_by_email(ADMIN_EMAIL)
        if admin_record:
            return admin_record["id"]
    return user_id

@app.get("/ref-db/diagnose")
def diagnose_ref_db():
    """Public diagnostic: check if reference_databases table exists."""
    result = {"table_exists": False, "row_count": None, "error": None,
              "users_table_ok": False, "supabase_ok": False}
    try:
        sb = get_supabase()
        result["supabase_ok"] = True
    except Exception as e:
        result["supabase_error"] = str(e)
        return result
    try:
        res = sb.table("reference_databases").select("id").limit(1).execute()
        result["table_exists"] = True
        result["row_count"] = len(res.data or [])
    except Exception as e:
        result["error"] = str(e)
    try:
        res2 = sb.table("users").select("id").limit(1).execute()
        result["users_table_ok"] = True
        result["users_count"] = len(res2.data or [])
    except Exception as e2:
        result["users_error"] = str(e2)
    return result

@app.get("/ref-db")
def list_ref_dbs(user: dict = Depends(get_current_user)):
    """List all reference databases uploaded by the current user."""
    sb = get_supabase()
    user_id = _get_user_db_id(user)
    try:
        res = sb.table("reference_databases").select("*").eq("user_id", user_id).order("created_at").execute()
        return {"files": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[ref-db list] DB error: {str(e)}")

@app.post("/ref-db/upload")
async def upload_ref_db(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload reference database, parse line count, upload to Supabase storage, cache locally, and store metadata in DB."""
    filename = file.filename
    if not filename.lower().endswith((".xlsx", ".xls", ".json")):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) and JSON (.json) files are accepted")
    
    # Enforce limit of max 5 files per user
    sb = get_supabase()
    user_id = _get_user_db_id(user)
    try:
        existing_res = sb.table("reference_databases").select("id").eq("user_id", user_id).execute()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"[ref-db upload] Table check failed — table may not exist: {str(e)}")
    if len(existing_res.data or []) >= 5:
        raise HTTPException(status_code=400, detail="Maximum limit of 5 reference database files reached. Please delete an existing file first.")
    
    content = await file.read()
    size_bytes = len(content)
    
    # Parse file to determine total line/QCM count
    line_count = 0
    try:
        import io
        if filename.lower().endswith(".json"):
            data = json.loads(content.decode("utf-8"))
            if isinstance(data, list):
                line_count = len(data)
            elif isinstance(data, dict):
                for key in ("qcms", "questions", "data", "items"):
                    if key in data and isinstance(data[key], list):
                        line_count = len(data[key])
                        break
                else:
                    line_count = len(data.keys())
        else:
            # Excel files: try pandas first, fall back to openpyxl
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(content))
                line_count = len(df)
            except Exception:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                count = sum(1 for r in rows if any(v is not None for v in r))
                line_count = max(0, count - 1)  # Subtract header row
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse reference database: {str(e)}")
    
    storage_path = f"{user['id']}/ref_dbs/{filename}"
    
    # Upload to Supabase Storage
    try:
        write_bytes_file(storage_path, content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload to storage: {str(e)}")
    
    db_record = {
        "user_id": user_id,
        "filename": filename,
        "storage_path": storage_path,
        "size_bytes": size_bytes,
        "line_count": line_count
    }
    
    try:
        # Delete if existing file with same name exists to avoid UNIQUE constraint violation
        sb.table("reference_databases").delete().eq("user_id", user_id).eq("filename", filename).execute()
        res = sb.table("reference_databases").insert(db_record).execute()
        record = res.data[0] if res.data else db_record
    except Exception as e:
        try:
            # Clean up uploaded storage file if database recording fails
            sb.storage.from_("qcm-projects").remove([storage_path])
        except:
            pass
        raise HTTPException(status_code=500, detail=f"Failed to save record to database: {str(e)}")
    
    # Cache locally on the container filesystem for execution compatibility
    try:
        local_dir = Path(f"/app/output/{user['id']}/ref_dbs")
        local_dir.mkdir(parents=True, exist_ok=True)
        (local_dir / filename).write_bytes(content)
    except Exception as e:
        print(f"[REF-DB] Local write failed: {e}")
    
    return record

@app.delete("/ref-db/{file_id}")
def delete_ref_db(file_id: str, user: dict = Depends(get_current_user)):
    """Delete reference database record from database, cloud storage, and local cache."""
    sb = get_supabase()
    
    user_id = _get_user_db_id(user)

    res = sb.table("reference_databases").select("*").eq("id", file_id).eq("user_id", user_id).limit(1).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Reference database file not found")
    
    record = res.data[0]
    filename = record["filename"]
    storage_path = record["storage_path"]
    
    # Delete from Database
    sb.table("reference_databases").delete().eq("id", file_id).execute()
    
    # Delete from Supabase Storage
    try:
        sb.storage.from_("qcm-projects").remove([storage_path])
    except Exception as e:
        print(f"[STORAGE] Failed to delete reference db from storage: {e}")
    
    # Delete local cached file
    try:
        local_file = Path(f"/app/output/{user['id']}/ref_dbs/{filename}")
        if local_file.exists():
            local_file.unlink()
    except Exception as e:
        print(f"[REF-DB] Failed to delete local file: {e}")
        
    return {"status": "deleted", "id": file_id}


@app.delete("/projects/{name}")
def delete_project(name: str, user: dict = Depends(get_current_user)):
    """Delete a project from Supabase Storage and local FS."""
    storage_prefix = f"{user['id']}/{name}"
    project_dir = Path(f"/app/output/{user['id']}/{name}")

    # Check existence in either location
    if not file_exists(f"{storage_prefix}/project.json") and not project_dir.exists():
        raise HTTPException(status_code=404, detail="Project not found")

    # Delete from Supabase Storage
    try:
        delete_prefix(storage_prefix)
    except Exception as e:
        print(f"[STORAGE] Supabase delete failed: {e}")

    # Delete from local filesystem
    if project_dir.exists():
        try:
            shutil.rmtree(project_dir)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to delete locally: {str(e)}")

    return {"deleted": name}

# --- Step Run History & Badges ---

_step_start_time: dict = {}

def _record_step_history(project: str, user_id: str, step_id: str, start_time: float, badge: str, stats: dict):
    storage_path = f"{user_id}/{project}/step_history.json"

    # Read existing history — try Supabase first, then local FS
    history = {}
    try:
        if file_exists(storage_path):
            history = json.loads(read_file(storage_path))
    except Exception:
        local_hist = Path(f"/app/output/{user_id}/{project}/step_history.json")
        try:
            history = json.loads(local_hist.read_text()) if local_hist.exists() else {}
        except Exception:
            history = {}

    entry = {
        "run_at": datetime.now().isoformat(),
        "badge": badge,
        "duration_seconds": round(time.time() - start_time, 1),
        **stats
    }
    if step_id not in history:
        history[step_id] = []
    history[step_id].append(entry)
    history_json = json.dumps(history, indent=2)

    # Write to Supabase Storage
    try:
        write_file(storage_path, history_json)
    except Exception as e:
        print(f"[STORAGE] step_history upload failed: {e}")

    # Also write locally (pipeline badge computation reads local FS)
    local_hist = Path(f"/app/output/{user_id}/{project}/step_history.json")
    local_hist.parent.mkdir(parents=True, exist_ok=True)
    local_hist.write_text(history_json)

def _compute_step_badge(project: str, user_id: str, step_id: str) -> tuple[str, dict]:
    stats = {}
    badge = "success"
    try:
        if step_id == "1":
            accepted = Path(f"/app/output/{user_id}/{project}/step1_extraction/accepted")
            rejected = Path(f"/app/output/{user_id}/{project}/step1_extraction/rejected")
            acc = len(list(accepted.glob("*.txt"))) if accepted.exists() else 0
            rej = len(list(rejected.glob("*.txt"))) if rejected.exists() else 0
            stats = {"pages_ok": acc, "pages_failed": rej}
            badge = "success" if rej == 0 and acc > 0 else ("warning" if acc > 0 else "error")
        elif step_id == "2":
            qcm_dir = Path(f"/app/output/{user_id}/{project}/step2_qcm")
            total_qcms = 0
            empty_pages = 0
            for f in (qcm_dir.rglob("*.json") if qcm_dir.exists() else []):
                try:
                    data = json.loads(f.read_text())
                    qcms = len(data) if isinstance(data, list) else 0
                    total_qcms += qcms
                    if qcms == 0: empty_pages += 1
                except: pass
            stats = {"qcms": total_qcms, "empty_pages": empty_pages}
            badge = "error" if total_qcms == 0 else ("warning" if empty_pages > 0 else "success")
        elif step_id == "5":
            merged = Path(f"/app/output/{user_id}/{project}/step5_json/merged_qcms.json")
            merged_count = len(json.loads(merged.read_text())) if merged.exists() else 0
            stats = {"merged_qcms": merged_count}
            badge = "error" if merged_count == 0 else "success"
        elif step_id in ("3", "4", "6", "7", "1.5", "1.6"):
            badge = "success"
        # step 8: no badge
    except Exception:
        badge = "success"
    return badge, stats

@app.get("/projects/{name}/step-history")
def get_step_history(name: str, user: dict = Depends(get_current_user)):
    # Try Supabase Storage first
    storage_path = f"{user['id']}/{name}/step_history.json"
    try:
        if file_exists(storage_path):
            return json.loads(read_file(storage_path))
    except Exception:
        pass
    # Fallback to local filesystem
    path = Path(f"/app/output/{user['id']}/{name}/step_history.json")
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except Exception:
        return {}

# --- Step Upload Helper (Group C fix) ---

def _upload_step_folder_to_storage(user_id: str, project: str, folder_name: str, storage_prefix: str) -> None:
    """
    Upload all files in a local step folder to Supabase Storage.
    Runs synchronously — must be called inside run_in_executor.
    Storage path: {user_id}/{project}/{folder_name}/{relative_file_path}
    """
    local_dir = Path(f"/app/output/{user_id}/{project}/{folder_name}")
    if not local_dir.exists():
        return
    uploaded = 0
    for f in local_dir.rglob("*"):
        if not f.is_file():
            continue
        rel = str(f.relative_to(local_dir)).replace("\\", "/")
        storage_path = f"{storage_prefix}/{rel}"
        try:
            write_bytes_file(storage_path, f.read_bytes())
            uploaded += 1
        except Exception as e:
            print(f"[STORAGE] upload failed {storage_path}: {e}")
    print(f"[STORAGE] uploaded {uploaded} file(s) for {storage_prefix}")


# --- Step Run Endpoint + Background Task ---

@app.post("/projects/{name}/steps/{step_id}/run")
async def run_step(name: str, step_id: str, body: dict, user: dict = Depends(get_current_user)):
    if job_manager.is_running(name, step_id):
        return {"error": "Step already running"}

    _apply_user_env(user)
    _step_start_time[f"{name}-{step_id}"] = time.time()

    # Pass the config body to the task
    task = asyncio.ensure_future(_run_step_task(name, user["id"], step_id, body))
    job_manager.set_running(name, step_id, task)
    return {"job_id": f"{name}-{step_id}-001"}

async def _run_step_task(project: str, user_id: str, step_id: str, config: dict):
    ctx_data = get_or_create(project, user_id)
    context = ctx_data["context"]
    tracker = ctx_data["tracker"]

    def log_callback(line: dict):
        job_manager.append_log(project, step_id, line)

    def _run_with_capture():
        with LogCapture(log_callback):
            _call_step(step_id, tracker, context, config)

    try:
        _STEP_FOLDER_MAP = {
            "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
            "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
            "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
        }
        folder_name = _STEP_FOLDER_MAP.get(step_id, f"step{step_id}")
        step_dir = Path(f"/app/output/{user_id}/{project}/{folder_name}")

        # Archive previous run to _history/ (local + Supabase)
        loop = asyncio.get_event_loop()
        if step_dir.exists() and any(step_dir.iterdir()):
            archive_ts = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
            archive_local = Path(f"/app/output/{user_id}/{project}/_history/step{step_id}/{archive_ts}")
            archive_storage = f"{user_id}/{project}/_history/step{step_id}/{archive_ts}"

            def _do_archive():
                archive_local.mkdir(parents=True, exist_ok=True)
                for f in step_dir.rglob("*"):
                    if f.is_file():
                        dest = archive_local / f.relative_to(step_dir)
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(str(f), str(dest))
                # Upload archive to Supabase
                _upload_step_folder_to_storage(user_id, project, f"_history/step{step_id}/{archive_ts}", archive_storage)

            await loop.run_in_executor(None, _do_archive)

        # Ensure source.pdf is on local disk before running any step.
        # After a container restart the local FS is wiped; download from Supabase.
        local_pdf = Path(f"/app/output/{user_id}/{project}/source.pdf")
        if not local_pdf.exists():
            storage_pdf = f"{user_id}/{project}/source.pdf"
            try:
                if file_exists(storage_pdf):
                    local_pdf.parent.mkdir(parents=True, exist_ok=True)
                    local_pdf.write_bytes(read_bytes_file(storage_pdf))
                    # Also fix project.json so pdf_path points to the local file
                    pjson_local = local_pdf.parent / "project.json"
                    pjson_local.write_text(json.dumps({"name": project, "pdf_path": str(local_pdf)}))
                    print(f"[RESTORE] ✅ PDF restored from Supabase to {local_pdf}")
                else:
                    print(f"[RESTORE] ⚠️ source.pdf not in Supabase at {storage_pdf}")
            except Exception as _e:
                print(f"[RESTORE] ❌ PDF download failed: {_e}")

        # Ensure Step 8 selected reference database is downloaded locally
        if step_id == "8" and config.get("ref_db_path"):
            ref_db_val = config.get("ref_db_path")
            # If it's a simple filename, retrieve it from Supabase storage and download locally
            if ref_db_val and not (ref_db_val.startswith("/") or ":" in ref_db_val or "\\" in ref_db_val):
                local_ref_dir = Path(f"/app/output/{user_id}/ref_dbs")
                local_ref_path = local_ref_dir / ref_db_val
                
                if not local_ref_path.exists():
                    storage_path = f"{user_id}/ref_dbs/{ref_db_val}"
                    try:
                        if file_exists(storage_path):
                            local_ref_dir.mkdir(parents=True, exist_ok=True)
                            local_ref_path.write_bytes(read_bytes_file(storage_path))
                            print(f"[RESTORE] ✅ Ref DB restored from Supabase to {local_ref_path}")
                        else:
                            print(f"[RESTORE] ⚠️ Ref DB not found in Supabase Storage at {storage_path}")
                    except Exception as ex:
                        print(f"[RESTORE] ❌ Ref DB download failed: {ex}")

        await loop.run_in_executor(None, _run_with_capture)

        job_manager.set_done(project, step_id)
        log_callback({"ts": datetime.now().strftime("%H:%M:%S"), "type": "ok", "text": f"\u2705 Step {step_id} completed successfully."})

        # All post-step I/O in one thread to keep event loop free
        def _do_post_step():
            # Save costs locally
            cost_path = f"/app/output/{user_id}/{project}/total_costs.json"
            Path(cost_path).parent.mkdir(parents=True, exist_ok=True)
            tracker.save(cost_path)
            # Upload costs to Supabase
            try:
                write_file(f"{user_id}/{project}/total_costs.json", Path(cost_path).read_text())
            except Exception as e:
                print(f"[STORAGE] cost upload failed: {e}")
            # Upload step output folder to Supabase
            step_storage_prefix = f"{user_id}/{project}/{folder_name}"
            _upload_step_folder_to_storage(user_id, project, folder_name, step_storage_prefix)

        await loop.run_in_executor(None, _do_post_step)

    except Exception as e:
        import traceback
        traceback.print_exc()
        job_manager.set_error(project, step_id)
        log_callback({"ts": datetime.now().strftime("%H:%M:%S"), "type": "error", "text": f"❌ Step {step_id} failed: {str(e)}"})

        try:
            _record_step_history(project, user_id, step_id, _step_start_time.get(f"{project}-{step_id}", time.time()), "error", {})
        except: pass

    else:
        try:
            badge, stats = _compute_step_badge(project, user_id, step_id)
            _record_step_history(project, user_id, step_id, _step_start_time.get(f"{project}-{step_id}", time.time()), badge, stats)
        except Exception as be:
            print(f"Badge recording failed: {be}")

def _call_step(step_id: str, tracker, context, config: dict):
    """Synchronous step dispatcher. Runs in a thread via run_in_executor."""
    
    def _auto_input(prompt=""):
        """Replace input() with automatic responses derived from UI config."""
        p = str(prompt).lower()
        print(f"[AUTO] {repr(prompt)}")
        
        # Step 1 — extraction method
        if "choice [1-2]" in p or "select method" in p:
            return "1" if config.get("method", "vision_ocr") == "vision_ocr" else "2"
        if "ocr guidance" in p or "your guidance" in p:
            return config.get("ocr_guidance", "")
        
        # Step 2 — page range
        if "your choice" in p or "page range" in p:
            return config.get("page_range", "all")
        
        # Step 2/6 — review prompts → always accept/skip
        if "[a]ccept" in p or "[r/c]" in p:
            return "a"
        if "[r]etry" in p or "[s]kip" in p or "debug" in p:
            return "s"
        
        # Step 3 — strategy choice
        if "choice:" in p:
            return ""
        
        # Step 6 — correction mode
        if "select mode [s/b]" in p:
            return "b" if config.get("correction_search_mode", "") == "all_pages" else "s"
        if "use these pages" in p:
            return "y"
        if "pages:" in p or "page number" in p:
            return config.get("correction_pages", "1")
        if "[u]se default" in p or "edit" in p:
            return "u"
        if "choice [1-4]" in p or "recovery" in p:
            return "1"
        if "correction" in p and "enter" in p:
            return ""
        
        # Step 8 — reference DB path prompt: return full resolved path so step8_matcher can find the file
        if prompt.strip() == ">":
            ref = config.get("ref_db_path", "")
            # If it's a plain filename, resolve to the full local container path
            if ref and not (ref.startswith("/") or ":" in ref or "\\" in ref):
                try:
                    uid = context.name.split("/")[0]
                except Exception:
                    uid = "admin"
                ref = f"/app/output/{uid}/ref_dbs/{ref}"
            return ref
        
        # Step 8 — interactive export → skip
        if "export custom xlsx" in p:
            return "n"
        if "use these settings" in p:
            return "y"
        if "select [1/2/3]" in p:
            return "1"
        if "threshold" in p:
            return ""
        if "weight" in p:
            return ""
        
        # Default — return empty (skip/use default)
        print(f"[AUTO-INPUT] Unhandled prompt, returning empty string")
        return ""

    # Install the patch
    builtins.input = _auto_input
    
    # Apply model overrides to environment before importing/running modules.
    # Only override when the UI explicitly sends a non-empty model string.
    if step_id == "1" and config.get("model"):
        os.environ["STEP1_MODEL"] = config["model"]
    if step_id == "2":
        if config.get("model_primary"):
            os.environ["STEP2_MODEL"] = config["model_primary"]
        if config.get("model_fallback"):
            os.environ["STEP2_FALLBACK_MODEL"] = config["model_fallback"]
    if step_id == "3":
        if config.get("model"):
            os.environ["STEP3_MODEL"] = config["model"]
        if config.get("model_fallback"):
            os.environ["STEP3_FALLBACK_MODEL"] = config["model_fallback"]
    if step_id == "6":
        if config.get("text_model"):
            os.environ["STEP6_TEXT_MODEL"] = config["text_model"]
        if config.get("all_pages_model"):
            os.environ["STEP6_ALL_PAGES_MODEL"] = config["all_pages_model"]
    
    if step_id == "8":
        if config.get("ref_db_path"):
            ref_db_val = config["ref_db_path"]
            # If it's a simple filename, resolve to the local user's folder path
            if ref_db_val and not (ref_db_val.startswith("/") or ":" in ref_db_val or "\\" in ref_db_val):
                uid = context.name.split("/")[0]
                ref_db_val = f"/app/output/{uid}/ref_dbs/{ref_db_val}"
            os.environ["REFERENCE_DB_PATH"] = ref_db_val
        if config.get("match_mode"):
            os.environ["MATCH_MODE"] = config["match_mode"]
        if config.get("threshold") is not None:
            os.environ["MATCH_THRESHOLD"] = str(config["threshold"])
        if config.get("text_weight") is not None:
            os.environ["MATCH_TEXT_WEIGHT"] = str(config["text_weight"])
        if config.get("corr_weight") is not None:
            os.environ["MATCH_CORRECTION_WEIGHT"] = str(config["corr_weight"])
        if config.get("color_green") is not None:
            os.environ["MATCH_COLOR_GREEN"] = str(config["color_green"])
        if config.get("color_yellow") is not None:
            os.environ["MATCH_COLOR_YELLOW"] = str(config["color_yellow"])

    # Late imports to avoid circular deps and only load what's needed
    from modules.step1_extraction import Step1Extraction
    from modules.step1_5_batch_text_fixer import Step1_5BatchTextFixer
    from modules.step1_6_intelligent_text_fixer import Step1_6IntelligentTextFixer
    from modules.step2_qcm_extract_batch import Step2QCMExtractBatch
    from modules.step3_metadata import Step3Metadata
    from modules.step4_format import Step4Format
    from modules.step5_builder import Step5Builder
    from modules.step6_corrections import Step6Corrections
    from modules.step7_categorization import Step7Categorization
    from modules.step8_matcher import Step8Matcher

    def _run_step4_auto(tracker, context, cfg):
        from modules.step4_format import Step4Format
        from modules.utils.template_library import TemplateLibrary
        name   = cfg.get("name") or "pediat"
        fields = cfg.get("fields", {})
        tmpl: dict = {}
        if fields.get("Num",          True):  tmpl["Num"]          = 0
        if fields.get("Text",         True):  tmpl["Text"]         = "Question text here..."
        if fields.get("Propositions", True):
            tmpl.update({"A":"Option A","B":"Option B","C":"Option C","D":"Option D","E":"Option E"})
        if fields.get("Correct",      True):  tmpl["Correct"]      = "ABC"
        if fields.get("Year",         True):  tmpl["Year"]         = "2024"
        if fields.get("Category",     True):  tmpl["categoryName"] = "Cardiologie"
        if fields.get("Subcategory",  False): tmpl["subcategoryName"] = "HTA"
        if fields.get("Source",       False): tmpl["Source"]       = "Alger"
        if fields.get("Tag",          True):  tmpl["Tag"]          = ["Alger", "2024"]
        if fields.get("ClinicalCase", False): tmpl["Cas"]          = "CAS CLINIQUE 1\r\nNarrative..."
        TemplateLibrary().save_template(name, tmpl)
        print(f"[STEP4] Template '{name}' saved: {list(tmpl.keys())}")
        Step4Format(tracker, context).run(auto_template=name)

    def _build_step6_config(cfg):
        source_map = {"ai_knowledge":"ai_knowledge","page_text":"page_text","auto_detect":"page_text","vision_ai":"vision"}
        backend_source = source_map.get(cfg.get("source","page_text"),"page_text")
        search_mode = "all_pages" if cfg.get("source")=="auto_detect" else cfg.get("correction_search_mode","all_pages")
        ai_mode = {"sequential":"S","batch":"B"}.get(cfg.get("ai_mode","sequential"),"S")
        return {
            "source": backend_source, "ai_mode": ai_mode,
            "correction_search_mode": search_mode,
            "pages": cfg.get("pages",""), "force_overwrite": cfg.get("force_overwrite",False),
            "vision": {"custom_prompt": cfg.get("vision_prompt","")},
            "page_text": {"extraction_guidance": cfg.get("page_text_guidance","")},
            "all_pages_scan": {"candidate_threshold": int(cfg.get("candidate_threshold",15)), "include_neighbors": bool(cfg.get("include_neighbors",True))},
        }

    step_map = {
        "1":   lambda: Step1Extraction(tracker, context).run(
                    # Always use the canonical local path — config.pdf_path may be
                    # stale (e.g. a Supabase storage key) after a container restart.
                    pdf_path=str(Path(f"/app/output/{context.name}/source.pdf")),
                    auto_ocr=(config.get("method", "") == "vision_ocr"),
                    ocr_guidance=config.get("ocr_guidance", "")
               ),
        "1.5": lambda: Step1_5BatchTextFixer(tracker, context).run(),
        "1.6": lambda: Step1_6IntelligentTextFixer(tracker, context).run(),
        "2":   lambda: Step2QCMExtractBatch(tracker, context).run(
                    page_range=config.get("page_range") or "all",
                    config={
                        "qcm_extraction": {
                            "extraction_guidance": config.get("extraction_guidance", ""),
                            "clinical_case_hints": config.get("clinical_case_hints", False),
                        }
                    }
               ),
        "3":   lambda: Step3Metadata(tracker, context).run(
                    auto_mode=True,
                    config=config.get("fields", {}),
                    global_pages=[
                        int(p.strip())
                        for p in str(config.get("global_pages", "1")).split(",")
                        if p.strip().isdigit()
                    ],
                ),
        "4":   lambda: _run_step4_auto(tracker, context, config),
        "5":   lambda: Step5Builder(tracker, context).run(),
        "6":   lambda: Step6Corrections(tracker, context).run(
                    pdf_path=config.get("pdf_path",""),
                    auto_mode=True,
                    config=_build_step6_config(config),
                ),
        "7":   lambda: Step7Categorization(tracker, context).run(),
        "8":   lambda: Step8Matcher(tracker, context).run(),
    }
    
    if step_id not in step_map:
        raise ValueError(f"Unknown step identifier: {step_id}")
    
    # Execute the step
    step_map[step_id]()

# --- Status + WebSocket Endpoints ---

@app.get("/projects/{name}/steps/{step_id}/status")
def get_step_status(name: str, step_id: str, user: dict = Depends(get_current_user)):
    return {
        "status": job_manager.get_status(name, step_id),
        "output_exists": step_output_exists(name, step_id, user["id"])
    }

@app.websocket("/ws/log/{project}/{step_id}")
async def ws_log(websocket: WebSocket, project: str, step_id: str):
    # WebSocket is intentionally left unprotected — Bearer header auth from the browser
    # is non-trivial over WS, and log data alone does not expose sensitive information.
    await websocket.accept()
    sent_count = 0
    try:
        while True:
            all_logs = job_manager.get_logs(project, step_id)
            # Push new lines
            while sent_count < len(all_logs):
                await websocket.send_text(json.dumps(all_logs[sent_count]))
                sent_count += 1
            
            # Exit loop if job is finished and all lines pushed
            status = job_manager.get_status(project, step_id)
            if status in ("done", "error") and sent_count >= len(all_logs):
                break
            
            await asyncio.sleep(0.3)
    except Exception:
        pass
    finally:
        try:
            await websocket.close()
        except:
            pass

# --- Cost Endpoints ---

@app.get("/projects/{name}/costs")
def get_project_costs(name: str, user: dict = Depends(get_current_user)):
    # Try Supabase Storage first
    storage_path = f"{user['id']}/{name}/total_costs.json"
    try:
        if file_exists(storage_path):
            data = json.loads(read_file(storage_path))
            return data.get("summary", data)
    except Exception:
        pass
    # Fallback to local filesystem
    cost_file = Path(f"/app/output/{user['id']}/{name}/total_costs.json")
    if cost_file.exists():
        try:
            data = json.loads(cost_file.read_text())
            return data.get("summary", data)
        except Exception:
            pass
    # Last resort: live in-memory tracker
    ctx_data = get_or_create(name, user["id"])
    return ctx_data["tracker"].get_total_summary()

@app.post("/projects/{name}/costs/save")
def save_project_costs(name: str, user: dict = Depends(get_current_user)):
    local_path = f"/app/output/{user['id']}/{name}/total_costs.json"
    Path(local_path).parent.mkdir(parents=True, exist_ok=True)
    ctx_data = get_or_create(name, user["id"])
    ctx_data["tracker"].save(local_path)
    # Also upload to Supabase Storage
    try:
        write_file(f"{user['id']}/{name}/total_costs.json", Path(local_path).read_text())
    except Exception as e:
        print(f"[STORAGE] costs/save upload failed: {e}")
    return {"saved_to": local_path}

@app.get("/costs/weekly")
def get_all_weekly_costs(user: dict = Depends(get_current_user)):
    return get_weekly_costs(user["id"])

# --- .env Endpoints ---

@app.get("/env")
def get_environment(user: dict = Depends(get_current_user)):
    env = read_env()
    result = {}
    for k, v in env.items():
        if k in EDITABLE_KEYS:
            result[k] = mask(v)
        elif any(x in k for x in ["MODEL", "ENABLE", "STEP"]):
            result[k] = v
    return result

@app.post("/env")
def update_environment(body: dict, user: dict = Depends(get_current_user)):
    updated = write_env_keys(body)
    # Update current process environment for immediate effect (hot-swap)
    env = read_env()
    for k in updated:
        os.environ[k] = env.get(k, "")
    return {"updated": updated}

@app.get("/env/raw")
def get_environment_raw(user: dict = Depends(get_current_user)):
    """Returns all env values unmasked for the settings UI."""
    env = read_env()
    result = {}
    for k, v in env.items():
        if k in EDITABLE_KEYS or any(x in k for x in ["MODEL", "ENABLE", "STEP"]):
            result[k] = v  # NO masking
    return result

ADMIN_ENV_PATH = Path("/app/admin.env")

@app.get("/admin/available-models")
def get_available_models(user: dict = Depends(get_current_user)):
    """
    Admin → return full model list from admin.env (unchanged).
    Regular user → return their allowed_models from users.json.
    If user's allowed_models is empty → fall back to admin.env global list.
    """
    # Parse admin.env (global list) — keep existing parsing logic
    global_models = {}
    if ADMIN_ENV_PATH.exists():
        for line in ADMIN_ENV_PATH.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, _, val = line.partition('=')
                global_models[key.strip().lower()] = [m.strip() for m in val.split(',') if m.strip()]

    # Admin always gets global list
    if user.get("is_admin"):
        return global_models

    # Regular user: use their assigned models if set, else fall back to global
    user_models = user.get("allowed_models", {})
    if user_models:
        # Merge: use user's list where set, fall back to global for missing slots
        merged = dict(global_models)
        merged.update(user_models)
        return merged

    return global_models

@app.get("/env/step-models")
def get_step_models(user: dict = Depends(get_current_user)):
    env = read_env()
    return {
        "step1":   {"primary": env.get("STEP1_MODEL"),   "fallback": env.get("STEP1_FALLBACK_MODEL")},
        "step1_5": {"primary": env.get("STEP1_5_MODEL"), "fallback": env.get("STEP1_5_FALLBACK_MODEL")},
        "step1_6": {"primary": env.get("STEP1_6_MODEL"), "fallback": env.get("STEP1_6_FALLBACK_MODEL")},
        "step2":   {"primary": env.get("STEP2_MODEL"),   "fallback": env.get("STEP2_FALLBACK_MODEL")},
        "step3":   {"primary": env.get("STEP3_MODEL"),   "fallback": env.get("STEP3_FALLBACK_MODEL")},
        "step6":   {
            "text_model":      env.get("STEP6_TEXT_MODEL"),
            "text_fallback":   env.get("STEP6_TEXT_FALLBACK_MODEL"),
            "all_pages_model": env.get("STEP6_ALL_PAGES_MODEL"),
            "ai_model":        env.get("STEP6_AI_MODEL")
        },
        "step7":   {"primary": env.get("STEP7_MODEL"),   "fallback": env.get("STEP7_FALLBACK_MODEL")},
        "step8":   {}
    }

# --- Batch Config + Template Endpoints ---

@app.get("/config/batch")
def get_batch_config(user: dict = Depends(get_current_user)):
    cfg_path = Path("/app/batch_config.yaml")
    if cfg_path.exists():
        try:
            with open(cfg_path) as f:
                return yaml.safe_load(f)
        except:
            pass
    return {}

@app.post("/config/batch")
def post_batch_config(body: dict, user: dict = Depends(get_current_user)):
    cfg_path = Path("/app/batch_config.yaml")
    # Monaco editor sends raw YAML string in body["yaml_content"]
    content = body.get("yaml_content", "")
    cfg_path.write_text(content)
    return {"saved": True}

@app.get("/templates")
def get_template_list(user: dict = Depends(get_current_user)):
    from modules.utils.template_library import TemplateLibrary
    return TemplateLibrary().list_templates()

@app.get("/projects/{name}/steps/{step_id}/output")
def get_step_output_files(name: str, step_id: str, user: dict = Depends(get_current_user)):
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    step_dir = Path(f"/app/output/{user['id']}/{name}/{folder_name}")

    # Local filesystem (during active container session)
    if step_dir.exists():
        files = []
        for f in sorted(step_dir.rglob("*")):
            if f.is_file():
                files.append({
                    "name": str(f.relative_to(step_dir)).replace("\\", "/"),
                    "size_bytes": f.stat().st_size,
                    "path": str(f).replace("\\", "/"),
                    "created_at": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                })
        if files:
            return {"files": files}

    # Fallback: list from Supabase Storage
    storage_prefix = f"{user['id']}/{name}/{folder_name}"
    try:
        items = list_files(storage_prefix)
        files = [
            {"name": it["name"], "size_bytes": it.get("metadata", {}).get("size", 0),
             "path": f"{storage_prefix}/{it['name']}", "created_at": ""}
            for it in items if it.get("id")
        ]
        return {"files": files}
    except Exception:
        return {"files": []}

@app.get("/projects/{name}/steps/{step_id}/output/{filename:path}")
def get_step_file_content(name: str, step_id: str, filename: str, user: dict = Depends(get_current_user)):
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    file_path = Path(f"/app/output/{user['id']}/{name}/{folder_name}") / filename

    ext = Path(filename).suffix.lower()

    # Try local first
    if file_path.exists() and file_path.is_file():
        if not str(file_path).startswith(f"/app/output/{user['id']}/{name}"):
            raise HTTPException(status_code=403, detail="Access denied")
        if ext in (".txt", ".json", ".yaml", ".yml", ".md"):
            return {"content": file_path.read_text(encoding="utf-8")}
        return {"binary": True, "size": file_path.stat().st_size}

    # Fallback: read from Supabase Storage
    storage_path = f"{user['id']}/{name}/{folder_name}/{filename}"
    try:
        if ext in (".txt", ".json", ".yaml", ".yml", ".md"):
            return {"content": read_file(storage_path)}
        data = read_bytes_file(storage_path)
        return {"binary": True, "size": len(data)}
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")


@app.get("/projects/{name}/steps/{step_id}/view/{filename:path}")
def view_step_file(name: str, step_id: str, filename: str, user: dict = Depends(get_current_user)):
    """Serve a file for inline viewing — local FS first, then Supabase signed URL."""
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    file_path = Path(f"/app/output/{user['id']}/{name}/{folder_name}") / filename
    if file_path.exists() and file_path.is_file():
        mime_type, _ = mimetypes.guess_type(str(file_path))
        return FileResponse(path=str(file_path), media_type=mime_type or "text/plain")
    # Fallback: Supabase signed URL
    try:
        url = get_signed_url(f"{user['id']}/{name}/{folder_name}/{filename}")
        return RedirectResponse(url)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

@app.get("/projects/{name}/steps/{step_id}/download/{filename:path}")
def download_step_file(name: str, step_id: str, filename: str, user: dict = Depends(get_current_user)):
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    file_path = Path(f"/app/output/{user['id']}/{name}/{folder_name}") / filename
    if file_path.exists() and file_path.is_file():
        mime_type, _ = mimetypes.guess_type(str(file_path))
        return FileResponse(path=str(file_path), media_type=mime_type or "application/octet-stream",
                            filename=file_path.name,
                            headers={"Content-Disposition": f'attachment; filename="{file_path.name}"'})
    # Fallback: Supabase signed URL
    try:
        url = get_signed_url(f"{user['id']}/{name}/{folder_name}/{filename}")
        return RedirectResponse(url)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

@app.get("/projects/{name}/steps/{step_id}/history")
def get_step_output_history(name: str, step_id: str, user: dict = Depends(get_current_user)):
    """List run snapshots — local FS first, then Supabase Storage."""
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    current_dir = Path(f"/app/output/{user['id']}/{name}/{folder_name}")
    runs = []

    # Current output — local
    if current_dir.exists():
        current_files = []
        for f in sorted(current_dir.rglob("*")):
            if f.is_file():
                current_files.append({
                    "name": str(f.relative_to(current_dir)).replace("\\", "/"),
                    "size_bytes": f.stat().st_size,
                    "created_at": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
        if current_files:
            runs.append({"run_id": "current", "label": "\u25b6 Current Output", "files": current_files})
    else:
        # Fallback: Supabase Storage for current
        try:
            items = list_files(f"{user['id']}/{name}/{folder_name}")
            current_files = [
                {"name": it["name"], "size_bytes": it.get("metadata", {}).get("size", 0), "created_at": ""}
                for it in items if it.get("id")
            ]
            if current_files:
                runs.append({"run_id": "current", "label": "\u25b6 Current Output", "files": current_files})
        except Exception:
            pass

    # History runs
    history_base = Path(f"/app/output/{user['id']}/{name}/_history/step{step_id}")
    if history_base.exists():
        for run_dir in sorted(history_base.iterdir(), reverse=True):
            if run_dir.is_dir():
                files = []
                for f in sorted(run_dir.rglob("*")):
                    if f.is_file():
                        files.append({
                            "name": str(f.relative_to(run_dir)).replace("\\", "/"),
                            "size_bytes": f.stat().st_size,
                            "created_at": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                        })
                runs.append({"run_id": run_dir.name, "label": run_dir.name, "files": files})
    else:
        # Fallback: Supabase Storage for history
        try:
            hist_prefix = f"{user['id']}/{name}/_history/step{step_id}"
            run_dirs = [it for it in list_files(hist_prefix) if not it.get("id")]  # folders have no id
            for rd in sorted(run_dirs, key=lambda x: x["name"], reverse=True):
                rname = rd["name"]
                items = list_files(f"{hist_prefix}/{rname}")
                files = [{"name": it["name"], "size_bytes": it.get("metadata", {}).get("size", 0), "created_at": ""}
                         for it in items if it.get("id")]
                runs.append({"run_id": rname, "label": rname, "files": files})
        except Exception:
            pass

    return {"runs": runs}

@app.get("/projects/{name}/steps/{step_id}/history/{run_id}/{filename:path}")
def get_history_file(name: str, step_id: str, run_id: str, filename: str, user: dict = Depends(get_current_user)):
    """Serve a history file — local FS first, then Supabase signed URL."""
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder_name = _SFMAP.get(step_id, f"step{step_id}")
    if run_id == "current":
        local_path = Path(f"/app/output/{user['id']}/{name}/{folder_name}") / filename
        storage_path = f"{user['id']}/{name}/{folder_name}/{filename}"
    else:
        local_path = Path(f"/app/output/{user['id']}/{name}/_history/step{step_id}/{run_id}") / filename
        storage_path = f"{user['id']}/{name}/_history/step{step_id}/{run_id}/{filename}"
    if local_path.exists():
        mime_type, _ = mimetypes.guess_type(str(local_path))
        return FileResponse(str(local_path), media_type=mime_type or "application/octet-stream")
    try:
        url = get_signed_url(storage_path)
        return RedirectResponse(url)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

@app.post("/projects/{name}/steps/{step_id}/open-sheets")
def open_in_google_sheets(name: str, step_id: str, body: dict, user: dict = Depends(get_current_user)):
    """Upload XLSX to Google Sheets — resolves file from local FS or Supabase Storage."""
    filename = body.get("filename", "")
    print(f"[SHEETS] Request: project={name}, step={step_id}, file={filename}, user={user['id']}")
    _SFMAP = {
        "1": "step1_extraction", "1.5": "step1_extraction", "1.6": "step1_extraction",
        "2": "step2_qcm", "3": "step3_metadata", "4": "step4_format",
        "5": "step5_json", "6": "step6_corrections", "7": "step7_categories", "8": "step8_matches",
    }
    folder = _SFMAP.get(step_id, f"step{step_id}")
    file_path = Path(f"/app/output/{user['id']}/{name}/{folder}") / filename

    # If not local, download from Supabase to a temp file
    if not file_path.exists():
        print(f"[SHEETS] File not local, downloading from Supabase...")
        import tempfile
        # Use raw path segments (supabase-py handles any required escaping internally)
        storage_path = f"{user['id']}/{name}/{folder}/{filename}"
        print(f"[SHEETS] Supabase path: {storage_path}")
        try:
            data = read_bytes_file(storage_path)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.write(data)
            tmp.close()
            file_path = Path(tmp.name)
            print(f"[SHEETS] Downloaded {len(data)} bytes to {file_path}")
        except Exception as e:
            print(f"[SHEETS] File not found in Supabase either: {e}")
            raise HTTPException(status_code=404, detail="File not found")

    else:
        print(f"[SHEETS] File found locally: {file_path} ({file_path.stat().st_size} bytes)")

    # Check Google client secret exists
    if not Path(GOOGLE_CLIENT_SECRET_PATH).exists():
        print(f"[SHEETS] ERROR: Google client secret not found at {GOOGLE_CLIENT_SECRET_PATH}")
        raise HTTPException(status_code=500, detail="Google client secret not configured on server")

    user_db_id = _get_user_db_id(user)
    creds = _get_google_creds(user_db_id)
    if not creds:
        print(f"[SHEETS] No Google creds for user {user['id']} (db_id={user_db_id}) — returning 401 NOT_AUTHORIZED")
        raise HTTPException(status_code=401, detail="NOT_AUTHORIZED")
    try:
        from googleapiclient.http import MediaFileUpload
        print(f"[SHEETS] Uploading to Google Drive...")
        drive_service = build("drive", "v3", credentials=creds)
        file_metadata = {"name": file_path.stem, "mimeType": "application/vnd.google-apps.spreadsheet"}
        media = MediaFileUpload(str(file_path),
                                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                resumable=True)
        uploaded = drive_service.files().create(body=file_metadata, media_body=media,
                                                fields="id,webViewLink").execute()
        sheets_url = uploaded.get("webViewLink")
        print(f"[SHEETS] Success! URL={sheets_url}")
        return {"url": sheets_url, "id": uploaded.get("id")}
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[SHEETS] Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Google Sheets upload failed: {str(e)}")

@app.post("/projects/{name}/step8/export-existing")
async def export_matches_only(name: str, body: dict, user: dict = Depends(get_current_user)):
    _apply_user_env(user)
    task = asyncio.ensure_future(_export_step8_task(name, user["id"], body))
    job_manager.set_running(name, "8-export", task)
    return {"job_id": f"{name}-8-export-001"}

async def _export_step8_task(project: str, email: str, body: dict):
    ctx_data = get_or_create(project, email)
    context = ctx_data["context"]
    tracker = ctx_data["tracker"]
    def log_callback(line: dict):
        job_manager.append_log(project, "8-export", line)
    def _run():
        with LogCapture(log_callback):
            from modules.step8_matcher import Step8Matcher
            matcher = Step8Matcher(tracker, context)
            # Override config from body
            if body.get("color_green"): matcher.color_green = float(body["color_green"])
            if body.get("color_yellow"): matcher.color_yellow = float(body["color_yellow"])
            # Patch input for the export prompts
            def _export_auto(prompt=""):
                p = str(prompt).lower()
                if "export custom xlsx" in p: return "y"
                if "select [1/2]" in p: return "1"  # range mode
                if "from %" in p: return str(int(float(body.get("export_from", 0)) * 100))
                if "to   %" in p or "to %" in p: return str(int(float(body.get("export_to", 0.6)) * 100))
                if "output filename" in p: return body.get("export_filename", "custom_export")
                return ""
            import builtins
            builtins.input = _export_auto
            matcher.export_from_existing()
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _run)
        job_manager.set_done(project, "8-export")
        log_callback({"ts": datetime.now().strftime("%H:%M:%S"), "type": "ok", "text": "✅ Custom export complete."})
    except Exception as e:
        job_manager.set_error(project, "8-export")
        log_callback({"ts": datetime.now().strftime("%H:%M:%S"), "type": "error", "text": f"❌ Export failed: {str(e)}"})

# --- Auto Run Logic ---

@app.post("/projects/{name}/autorun")
async def start_autorun_sequence(name: str, body: dict, user: dict = Depends(get_current_user)):
    _apply_user_env(user)
    # Sequential execution of multiple steps
    asyncio.ensure_future(_autorun_task(name, user["id"], body))
    return {"job_id": f"{name}-autorun-001"}

async def _autorun_task(project: str, email: str, body: dict):
    start = str(body.get("start_step", "1"))
    end   = str(body.get("end_step",   "7"))
    run_config = body.get("run_config", {})
    
    # Valid step sequence
    sequence = ["1", "1.5", "1.6", "2", "3", "4", "5", "6", "7", "8"]
    
    # Filter sequence by start/end constraints
    try:
        start_idx = sequence.index(str(start))
        end_idx = sequence.index(str(end))
        active_sequence = sequence[start_idx:end_idx+1]
    except ValueError:
        return # Silent fail for bad range

    for step_id in active_sequence:
        cfg = run_config.get(f"step{step_id}", {})
        # Re-use the step task runner logic
        await _run_step_task(project, email, step_id, cfg)
        
        # Stop sequence if any step fails
        if job_manager.get_status(project, step_id) == "error":
            break
