"""
step8_matcher.py
Step 8: Smart QCM Similarity Matcher

Compares QCMs from the current project against a reference database.
- Auto-detects the latest available step output (step7 > step6 > step5)
- Reference DB path is set in .env (REFERENCE_DB_PATH) — supports .xlsx and .json
- Three similarity modes: text_only | full | weighted
- All thresholds and color bands are configurable via .env
- Outputs: step8_matches.json + step8_matches.xlsx (color-coded report)
"""

import json
import uuid
from datetime import datetime
import os
import re
import sys
import unicodedata
import multiprocessing
from multiprocessing import Pool, cpu_count
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional

import numpy as np

try:
    from rapidfuzz import fuzz, process
    from tqdm import tqdm
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from modules.utils.cost_tracker import CostTracker


# ── Step output priority (latest first) ──────────────────────────────────────
STEP_SOURCES = [
    ("step7_categories", "final_qcms.json",      "Step 7 (Categorized)"),
    ("step6_corrections","corrected_qcms.json",   "Step 6 (Corrected)"),
    ("step5_json",       "merged_qcms.json",      "Step 5 (Merged)"),
]

# ── Report column layout ──────────────────────────────────────────────────────
REPORT_COLUMNS = [
    ("Source #",         14),
    ("Source Question",  60),
    ("Source Correct",   14),
    ("Source Year",      10),
    ("Source Sub",       28),
    ("Match Score",      12),
    ("Text Score",       12),
    ("Corr Score",       12),
    ("Mode",             12),
    ("Ref #",            10),
    ("Ref Question",     60),
    ("Ref Correct",      14),
    ("Ref Year",         10),
    ("Ref Sub",          28),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKER (must be module-level for multiprocessing pickling)
# ═══════════════════════════════════════════════════════════════════════════════

def _match_chunk_worker(args):
    """
    Multiprocessing worker: matches a chunk of source QCMs against all ref texts.
    args = (src_indices, src_texts_chunk, ref_texts, threshold_pct)
    Returns list of (src_idx, [(ref_idx, score), ...])
    """
    src_indices, src_texts_chunk, ref_texts, threshold_pct = args
    results = []
    for i, query_text in zip(src_indices, src_texts_chunk):
        if not query_text or len(query_text) < 5:
            results.append((i, []))
            continue
        matches = process.extract(
            query_text,
            ref_texts,
            scorer=fuzz.ratio,
            score_cutoff=threshold_pct,
            limit=5,         # Top 5 candidates per QCM
        )
        clean = sorted(
            [(m[2], m[1] / 100.0) for m in matches],
            key=lambda x: x[1], reverse=True
        )
        results.append((i, clean))
    return results


def _match_chunk_worker_weighted(args):
    """
    Weighted mode worker: matches separately on text part and correction part.
    args = (src_indices, src_text_chunk, src_corr_chunk,
            ref_texts, ref_corrs, threshold_pct, text_w, corr_w)
    """
    (src_indices, src_text_chunk, src_corr_chunk,
     ref_texts, ref_corrs, threshold_pct, text_w, corr_w) = args

    results = []
    for i, t_query, c_query in zip(src_indices, src_text_chunk, src_corr_chunk):
        if not t_query or len(t_query) < 5:
            results.append((i, []))
            continue

        # Text score candidates (broad, no cutoff yet)
        raw = process.extract(
            t_query, ref_texts,
            scorer=fuzz.ratio, score_cutoff=0, limit=20
        )

        combined = []
        for m in raw:
            ref_idx = m[2]
            t_score = m[1] / 100.0
            # Correction score (0 if either side empty)
            if c_query and ref_corrs[ref_idx]:
                c_score = fuzz.ratio(c_query, ref_corrs[ref_idx]) / 100.0
            else:
                c_score = 0.0
            final = text_w * t_score + corr_w * c_score
            if final >= threshold_pct / 100.0:
                combined.append((ref_idx, final, t_score, c_score))

        combined.sort(key=lambda x: x[1], reverse=True)
        results.append((i, combined[:5]))
    return results


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class Step8Matcher:
    """Step 8: Compare project QCMs against a reference DB for similarity."""

    def __init__(self, cost_tracker: CostTracker, project_context=None):
        self.cost_tracker = cost_tracker
        self.context = project_context

        # Read .env config (with sane defaults)
        self.ref_db_path   = os.getenv("REFERENCE_DB_PATH", "").strip().strip('"').strip("'")
        self.threshold     = float(os.getenv("MATCH_THRESHOLD", "0.75"))
        self.mode          = os.getenv("MATCH_MODE", "text_only").strip().lower()
        self.text_weight   = float(os.getenv("MATCH_TEXT_WEIGHT", "0.7"))
        self.corr_weight   = float(os.getenv("MATCH_CORRECTION_WEIGHT", "0.3"))
        self.color_green   = float(os.getenv("MATCH_COLOR_GREEN", "0.90"))
        self.color_yellow  = float(os.getenv("MATCH_COLOR_YELLOW", "0.75"))
        self._source_label = "Unknown"  # set during _resolve_source

    # ── Public entry point ────────────────────────────────────────────────────

    def run(self) -> Dict:
        """Main execution for Step 8."""
        print("\n" + "=" * 60)
        print("STEP 8: QCM SIMILARITY MATCHER")
        print("=" * 60)

        if not HAS_RAPIDFUZZ:
            print("❌ rapidfuzz not installed. Run: pip install rapidfuzz tqdm")
            return {}
        if not HAS_OPENPYXL:
            print("❌ openpyxl not installed. Run: pip install openpyxl")
            return {}

        # 1. Resolve source QCMs
        source_path, source_label = self._resolve_source()
        if not source_path:
            return {}

        with open(source_path, "r", encoding="utf-8") as f:
            source_qcms: List[Dict] = json.load(f)
        print(f"📄 Source: {source_label}  ({len(source_qcms)} QCMs)")

        # 2. Resolve + load reference DB
        ref_path = self._resolve_ref_db()
        if not ref_path:
            return {}

        ref_qcms = self._load_reference_db(ref_path)
        if not ref_qcms:
            print("❌ Reference DB is empty or could not be parsed.")
            return {}
        print(f"📚 Reference: {Path(ref_path).name}  ({len(ref_qcms)} QCMs)")

        # 3. Interactive config (confirm or override .env values)
        self._configure_interactive()

        # 4. Run matching
        print(f"\n⚙️  Mode={self.mode}  |  Threshold={self.threshold:.0%}")
        if self.mode == "weighted":
            print(f"   Weights: text={self.text_weight:.0%}  correction={self.corr_weight:.0%}")

        match_records = self._run_matching(source_qcms, ref_qcms)

        if not match_records:
            print("\n⚠️  No matches found above the threshold.")
            return {"total_matches": 0}

        # 5. Save results
        output_dir = self._get_output_dir()
        self._save_json(match_records, output_dir / "step8_matches.json")
        # Count only actually-matched records (similarity > 0) for summary
        matched_only = [r for r in match_records if r["best_match"]["similarity"] > 0]
        self._save_summary(matched_only, output_dir / "step8_summary.json")
        self._save_xlsx_report(matched_only, output_dir / "step8_matches.xlsx")
        self._print_summary(matched_only, source_qcms)

        # 6. Offer custom export
        self._offer_custom_export(match_records, output_dir)

        return {"total_matches": len(matched_only), "output_dir": str(output_dir)}

    # ── Source resolution ─────────────────────────────────────────────────────

    def _resolve_source(self) -> Tuple[Optional[Path], str]:
        """Auto-detect latest available step output (step7 > step6 > step5)."""
        for step_folder, filename, label in STEP_SOURCES:
            if self.context:
                candidate = self.context.base_path / step_folder / filename
            else:
                candidate = Path("output") / step_folder / filename

            if candidate.exists():
                print(f"🔍 Auto-detected source: {label}  →  {candidate.name}")
                self._source_label = label
                return candidate, label

        print("❌ No source QCMs found.")
        print("   Please run at least Step 5 first.")
        return None, ""

    # ── Reference DB ──────────────────────────────────────────────────────────

    def _resolve_ref_db(self) -> Optional[str]:
        """Get reference DB path from .env or ask interactively."""
        if self.ref_db_path and Path(self.ref_db_path).exists():
            return self.ref_db_path

        if self.ref_db_path:
            print(f"⚠️  REFERENCE_DB_PATH in .env not found: {self.ref_db_path}")

        print("\n📂 Enter path to reference database (.xlsx or .json):")
        path = input("> ").strip().strip('"').strip("'")
        if not path or not Path(path).exists():
            print("❌ File not found.")
            return None
        return path

    def _load_reference_db(self, path: str) -> List[Dict]:
        """Load reference DB from .xlsx or .json into a list of QCM dicts."""
        p = Path(path)
        suffix = p.suffix.lower()

        if suffix == ".json":
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return data
            # Some JSON DBs wrap in {qcms: [...]}
            for key in ("qcms", "questions", "data", "items"):
                if key in data and isinstance(data[key], list):
                    return data[key]
            print("⚠️  Unexpected JSON structure — trying root object as list.")
            return list(data.values()) if isinstance(data, dict) else []

        elif suffix in (".xlsx", ".xls"):
            if not HAS_PANDAS:
                # Fallback: read with openpyxl directly
                return self._load_xlsx_openpyxl(p)
            try:
                df = pd.read_excel(p, dtype=str)
                df = df.where(df.notna(), None)
                return df.to_dict(orient="records")
            except Exception as e:
                print(f"⚠️  pandas failed ({e}), trying openpyxl fallback...")
                return self._load_xlsx_openpyxl(p)
        else:
            print(f"❌ Unsupported file format: {suffix}  (use .xlsx or .json)")
            return []

    def _load_xlsx_openpyxl(self, path: Path) -> List[Dict]:
        """Read xlsx without pandas using openpyxl."""
        from openpyxl import load_workbook
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(rows[0])]
            result = []
            for row in rows[1:]:
                if all(v is None for v in row):
                    continue
                result.append({
                    headers[i]: (str(v).strip() if v is not None else "")
                    for i, v in enumerate(row)
                })
            return result
        except Exception as e:
            print(f"❌ Could not read XLSX: {e}")
            return []

    # ── Interactive configuration ──────────────────────────────────────────────

    def _configure_interactive(self):
        """Let user adjust mode/threshold or accept .env defaults."""
        print(f"\n⚙️  Current configuration (.env defaults):")
        print(f"   Mode      : {self.mode}")
        print(f"   Threshold : {self.threshold:.0%}")
        if self.mode == "weighted":
            print(f"   Text W    : {self.text_weight:.0%}")
            print(f"   Corr W    : {self.corr_weight:.0%}")

        ans = input("\nUse these settings? [Y/n]: ").strip().lower()
        if ans in ("n", "no"):
            self._prompt_mode()
            self._prompt_threshold()
            if self.mode == "weighted":
                self._prompt_weights()

    def _prompt_mode(self):
        print("\n  Similarity mode:")
        print("   1. text_only  — question + options only (ignore Correct)")
        print("   2. full       — include Correct field")
        print("   3. weighted   — separate text + correction scores")
        c = input("  Select [1/2/3, default=1]: ").strip()
        self.mode = {"2": "full", "3": "weighted"}.get(c, "text_only")

    def _prompt_threshold(self):
        raw = input(f"\n  Threshold (0.0–1.0) [current={self.threshold:.2f}]: ").strip()
        if raw:
            try:
                self.threshold = max(0.0, min(1.0, float(raw)))
            except ValueError:
                pass
        print(f"  ✓ Threshold: {self.threshold:.0%}")

    def _prompt_weights(self):
        tw = input(f"  Text weight [current={self.text_weight:.2f}]: ").strip()
        if tw:
            try:
                self.text_weight = float(tw)
                self.corr_weight = round(1.0 - self.text_weight, 4)
                print(f"  ✓ text={self.text_weight:.0%}  correction={self.corr_weight:.0%}")
            except ValueError:
                pass

    # ── Text normalization ─────────────────────────────────────────────────────

    @staticmethod
    def _normalize(s: Any) -> str:
        if s is None or (isinstance(s, float) and np.isnan(s)):
            return ""
        s = str(s)
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.lower()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def _build_text(self, qcm: Dict) -> str:
        """Build normalized text+options string (no Correct)."""
        parts = [self._normalize(qcm.get("Text") or qcm.get("text", ""))]
        for k in ("A", "B", "C", "D", "E"):
            v = qcm.get(k)
            if v:
                parts.append(self._normalize(v))
        return " ".join(p for p in parts if p)

    def _build_full(self, qcm: Dict) -> str:
        """Normalized text + options + Correct field."""
        base = self._build_text(qcm)
        correct = self._normalize(qcm.get("Correct") or qcm.get("correct", ""))
        return (base + " " + correct).strip()

    def _build_correction(self, qcm: Dict) -> str:
        """Normalized Correct field only (for weighted mode)."""
        return self._normalize(qcm.get("Correct") or qcm.get("correct", ""))

    # ── Matching engine ───────────────────────────────────────────────────────

    def _run_matching(self, source_qcms: List[Dict], ref_qcms: List[Dict]) -> List[Dict]:
        """Dispatch to the right matching engine based on mode."""
        if self.mode == "weighted":
            return self._match_weighted(source_qcms, ref_qcms)
        else:
            return self._match_standard(source_qcms, ref_qcms)

    def _match_standard(self, source_qcms: List[Dict], ref_qcms: List[Dict]) -> List[Dict]:
        """text_only or full mode — single normalized string per QCM."""
        build_fn = self._build_full if self.mode == "full" else self._build_text
        src_texts = [build_fn(q) for q in source_qcms]
        ref_texts = [build_fn(q) for q in ref_qcms]

        n_workers  = cpu_count()
        chunk_size = max(1, len(source_qcms) // n_workers)
        indices    = list(range(len(source_qcms)))

        chunks = []
        for i in range(0, len(source_qcms), chunk_size):
            idx_chunk  = indices[i: i + chunk_size]
            text_chunk = src_texts[i: i + chunk_size]
            chunks.append((idx_chunk, text_chunk, ref_texts, self.threshold * 100))

        print(f"🚀 Matching using {n_workers} CPU cores...")
        raw_results = []
        with Pool(processes=n_workers) as pool:
            for res in tqdm(
                pool.imap_unordered(_match_chunk_worker, chunks),
                total=len(chunks), unit="chunk"
            ):
                raw_results.extend(res)

        raw_results.sort(key=lambda x: x[0])

        # Build lookup: source_index -> (best_ref_idx, best_score, candidates)
        matched_lookup: Dict[int, Any] = {}
        for src_idx, candidates in raw_results:
            if candidates:
                matched_lookup[src_idx] = candidates

        records = []
        for src_idx, src_qcm in enumerate(source_qcms):
            candidates = matched_lookup.get(src_idx, [])
            qcm_id = str(uuid.uuid4())

            if candidates:
                best_ref_idx, best_score = candidates[0]
                ref_qcm = ref_qcms[best_ref_idx]
                best_match = {
                    "ref_index":        best_ref_idx,
                    "similarity":       round(best_score, 4),
                    "text_similarity":  round(best_score, 4),
                    "corr_similarity":  None,
                    "mode":             self.mode,
                    "ref_qcm":          self._slim(ref_qcm),
                }
                all_cands = [
                    {"ref_index": ri, "similarity": round(sc, 4)}
                    for ri, sc in candidates
                ]
            else:
                best_match = {
                    "ref_index":        None,
                    "similarity":       0.0,
                    "text_similarity":  0.0,
                    "corr_similarity":  None,
                    "mode":             self.mode,
                    "ref_qcm":          {},
                }
                all_cands = []

            records.append({
                "qcm_id":        qcm_id,
                "source_step":   self._source_label,
                "source_index":  src_idx,
                "source_qcm":    self._slim(src_qcm),
                "source_qcm_full": self._full_source(src_qcm),
                "best_match":    best_match,
                "all_candidates": all_cands,
            })

        # Sort by descending similarity (unmatched go to bottom)
        records.sort(key=lambda r: r["best_match"]["similarity"], reverse=True)
        return records

    def _match_weighted(self, source_qcms: List[Dict], ref_qcms: List[Dict]) -> List[Dict]:
        """Weighted mode — combines text and correction scores."""
        src_texts = [self._build_text(q) for q in source_qcms]
        src_corrs = [self._build_correction(q) for q in source_qcms]
        ref_texts = [self._build_text(q) for q in ref_qcms]
        ref_corrs = [self._build_correction(q) for q in ref_qcms]

        n_workers  = cpu_count()
        chunk_size = max(1, len(source_qcms) // n_workers)
        indices    = list(range(len(source_qcms)))

        chunks = []
        for i in range(0, len(source_qcms), chunk_size):
            idx_chunk  = indices[i: i + chunk_size]
            text_chunk = src_texts[i: i + chunk_size]
            corr_chunk = src_corrs[i: i + chunk_size]
            chunks.append((
                idx_chunk, text_chunk, corr_chunk,
                ref_texts, ref_corrs,
                self.threshold * 100,
                self.text_weight, self.corr_weight
            ))

        print(f"🚀 Weighted matching using {n_workers} CPU cores...")
        raw_results = []
        with Pool(processes=n_workers) as pool:
            for res in tqdm(
                pool.imap_unordered(_match_chunk_worker_weighted, chunks),
                total=len(chunks), unit="chunk"
            ):
                raw_results.extend(res)

        raw_results.sort(key=lambda x: x[0])

        # Build lookup: source_index -> candidates
        matched_lookup: Dict[int, Any] = {}
        for src_idx, candidates in raw_results:
            if candidates:
                matched_lookup[src_idx] = candidates

        records = []
        for src_idx, src_qcm in enumerate(source_qcms):
            candidates = matched_lookup.get(src_idx, [])
            qcm_id = str(uuid.uuid4())

            if candidates:
                best = candidates[0]   # (ref_idx, final, t_score, c_score)
                ref_qcm = ref_qcms[best[0]]
                best_match = {
                    "ref_index":        best[0],
                    "similarity":       round(best[1], 4),
                    "text_similarity":  round(best[2], 4),
                    "corr_similarity":  round(best[3], 4),
                    "mode":             "weighted",
                    "ref_qcm":          self._slim(ref_qcm),
                }
                all_cands = [
                    {"ref_index": ri, "similarity": round(sc, 4),
                     "text_similarity": round(ts, 4),
                     "corr_similarity": round(cs, 4)}
                    for ri, sc, ts, cs in candidates
                ]
            else:
                best_match = {
                    "ref_index":        None,
                    "similarity":       0.0,
                    "text_similarity":  0.0,
                    "corr_similarity":  0.0,
                    "mode":             "weighted",
                    "ref_qcm":          {},
                }
                all_cands = []

            records.append({
                "qcm_id":         qcm_id,
                "source_step":    self._source_label,
                "source_index":   src_idx,
                "source_qcm":     self._slim(src_qcm),
                "source_qcm_full": self._full_source(src_qcm),
                "best_match":     best_match,
                "all_candidates": all_cands,
            })

        records.sort(key=lambda r: r["best_match"]["similarity"], reverse=True)
        return records

    # ── Output helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _slim(qcm: Dict) -> Dict:
        """Keep only essential display fields in match records."""
        keys = ["Num", "Text", "A", "B", "C", "D", "E",
                "Correct", "Year", "subcategoryName", "categoryName",
                "Tag", "Source"]
        return {k: qcm.get(k, qcm.get(k.lower(), "")) for k in keys}

    @staticmethod
    def _full_source(qcm: Dict) -> Dict:
        """Keep all source QCM fields for export (no filtering)."""
        result = dict(qcm)
        # Normalize Tag to string for storage
        tag = result.get("Tag", "")
        if isinstance(tag, list):
            result["Tag"] = ", ".join(str(t) for t in tag)
        return result

    def _get_output_dir(self) -> Path:
        if self.context:
            d = self.context.base_path / "step8_matches"
        else:
            d = Path("output/step8_matches")
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _save_json(self, records: List[Dict], path: Path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2, default=str)
        print(f"💾 JSON  → {path.name}")

    def _save_summary(self, records: List[Dict], path: Path):
        scores = [r["best_match"]["similarity"] for r in records]
        summary = {
            "total_matches":    len(records),
            "avg_similarity":   round(sum(scores) / len(scores), 4) if scores else 0,
            "green_matches":    sum(1 for s in scores if s >= self.color_green),
            "yellow_matches":   sum(1 for s in scores if self.color_yellow <= s < self.color_green),
            "red_matches":      sum(1 for s in scores if s < self.color_yellow),
            "threshold":        self.threshold,
            "mode":             self.mode,
            "color_green":      self.color_green,
            "color_yellow":     self.color_yellow,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)
        print(f"📊 Summary → {path.name}")

    def _save_xlsx_report(self, records: List[Dict], path: Path):
        """Save a color-coded XLSX match report."""
        if not HAS_OPENPYXL:
            print("⚠️  openpyxl missing — skipping XLSX report")
            return

        # Color fills
        fill_green  = PatternFill("solid", fgColor="C6EFCE")
        fill_yellow = PatternFill("solid", fgColor="FFEB9C")
        fill_red    = PatternFill("solid", fgColor="FFC7CE")
        fill_header = PatternFill("solid", fgColor="1F4E79")

        font_header  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        font_regular = Font(name="Calibri", size=10)
        border = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"),
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Matches"

        # ── Header ──
        col_names  = [c[0] for c in REPORT_COLUMNS]
        col_widths = [c[1] for c in REPORT_COLUMNS]
        for ci, (name, width) in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font   = font_header
            cell.fill   = fill_header
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # ── Data rows ──
        for ri, rec in enumerate(records, start=2):
            src  = rec["source_qcm"]
            bm   = rec["best_match"]
            ref  = bm["ref_qcm"]
            score     = bm["similarity"]
            t_score   = bm.get("text_similarity")
            c_score   = bm.get("corr_similarity")

            # Choose row fill based on score
            if score >= self.color_green:
                row_fill = fill_green
            elif score >= self.color_yellow:
                row_fill = fill_yellow
            else:
                row_fill = fill_red

            row_values = [
                src.get("Num", ""),
                src.get("Text", ""),
                src.get("Correct", ""),
                src.get("Year", ""),
                src.get("subcategoryName", ""),
                f"{score:.1%}",
                f"{t_score:.1%}" if t_score is not None else "",
                f"{c_score:.1%}" if c_score is not None else "",
                bm.get("mode", self.mode),
                ref.get("Num", ""),
                ref.get("Text", ""),
                ref.get("Correct", ""),
                ref.get("Year", ""),
                ref.get("subcategoryName", ""),
            ]

            for ci, val in enumerate(row_values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill   = row_fill
                cell.font   = font_regular
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(ci in (2, 11))  # wrap Question Text columns
                )

        # ── Legend sheet ──
        ls = wb.create_sheet("Legend")
        ls.column_dimensions["A"].width = 20
        ls.column_dimensions["B"].width = 50
        legend_rows = [
            ("Color",    "Meaning"),
            ("🟢 Green",  f"Similarity ≥ {self.color_green:.0%}  (high confidence match)"),
            ("🟡 Yellow", f"Similarity {self.color_yellow:.0%} – {self.color_green:.0%}  (moderate match)"),
            ("🔴 Red",    f"Similarity < {self.color_yellow:.0%}  (weak match, review manually)"),
        ]
        for row_data in legend_rows:
            ls.append(list(row_data))

        # ── Auto-filter ──
        ws.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}1"

        try:
            wb.save(path)
            print(f"📊 XLSX  → {path.name}")
        except Exception as e:
            print(f"⚠️  Could not save XLSX report: {e}")

    def _print_summary(self, records: List[Dict], source_qcms: List[Dict]):
        scores = [r["best_match"]["similarity"] for r in records]
        green  = sum(1 for s in scores if s >= self.color_green)
        yellow = sum(1 for s in scores if self.color_yellow <= s < self.color_green)
        red    = sum(1 for s in scores if s < self.color_yellow)

        print("\n" + "=" * 60)
        print(f"✅ Step 8 Complete  —  {len(records)}/{len(source_qcms)} QCMs matched")
        print(f"   🟢 High   (≥{self.color_green:.0%}) : {green}")
        print(f"   🟡 Medium ({self.color_yellow:.0%}–{self.color_green:.0%}) : {yellow}")
        print(f"   🔴 Low    (<{self.color_yellow:.0%}) : {red}")
        if scores:
            print(f"   📈 Average similarity : {sum(scores)/len(scores):.1%}")
        print("=" * 60)

    # ── Internal label — updated by _resolve_source ───────────────────────────
    # (Instance attribute _source_label is set in __init__)

    # ═══════════════════════════════════════════════════════════════════════
    # CUSTOM EXPORT FEATURE
    # ═══════════════════════════════════════════════════════════════════════

    def _offer_custom_export(self, all_records: List[Dict], output_dir: Path):
        """Interactive menu to export a filtered custom XLSX after Step 8 runs."""
        if not HAS_OPENPYXL:
            return

        print("\n" + "-" * 60)
        ans = input("📤 Export custom XLSX? [Y/n]: ").strip().lower()
        if ans in ("n", "no"):
            return

        print("\n  Export mode:")
        print("    1. Similarity range  (e.g. all QCMs with 0%–60% similarity)")
        print("    2. QCM ID list       (e.g. specific QCMs by their ID)")
        mode = input("  Select [1/2]: ").strip()

        if mode == "2":
            raw_ids = input("  Enter QCM IDs (comma-separated): ").strip()
            selected_ids = set(i.strip() for i in raw_ids.split(",") if i.strip())
            filtered = [r for r in all_records if r.get("qcm_id") in selected_ids]
            if not filtered:
                print("  ⚠️  No QCMs found with those IDs.")
                return
            label = f"IDs({len(filtered)})"
        else:
            # Range mode
            try:
                from_raw = input("  From % [0–100, default=0]: ").strip()
                to_raw   = input("  To   % [0–100, default=60]: ").strip()
                from_pct = float(from_raw) / 100.0 if from_raw else 0.0
                to_pct   = float(to_raw)   / 100.0 if to_raw   else 0.60
            except ValueError:
                print("  ❌ Invalid input. Using 0–60%.")
                from_pct, to_pct = 0.0, 0.60

            filtered = [
                r for r in all_records
                if from_pct <= r["best_match"]["similarity"] <= to_pct
            ]
            if not filtered:
                print(f"  ⚠️  No QCMs in range {from_pct:.0%}–{to_pct:.0%}.")
                return
            label = f"{from_pct:.0%}to{to_pct:.0%}"

        # Output filename
        fname_raw = input(f"  Output filename [default: custom_{label}]: ").strip()
        fname = fname_raw if fname_raw else f"custom_{label}"
        if not fname.endswith(".xlsx"):
            fname += ".xlsx"
        out_path = output_dir / fname

        self._save_custom_xlsx(filtered, out_path)

    def _save_custom_xlsx(self, records: List[Dict], path: Path):
        """Generate a full custom XLSX export with all source QCM fields + similarity + ref."""
        if not HAS_OPENPYXL:
            print("⚠️  openpyxl missing — cannot save XLSX")
            return

        # Color fills
        fill_green  = PatternFill("solid", fgColor="C6EFCE")
        fill_yellow = PatternFill("solid", fgColor="FFEB9C")
        fill_red    = PatternFill("solid", fgColor="FFC7CE")
        fill_grey   = PatternFill("solid", fgColor="DDDDDD")   # unmatched
        fill_header = PatternFill("solid", fgColor="1F4E79")

        font_header  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        font_regular = Font(name="Calibri", size=10)
        border = Border(
            left=Side(style="thin", color="AAAAAA"),
            right=Side(style="thin", color="AAAAAA"),
            top=Side(style="thin", color="AAAAAA"),
            bottom=Side(style="thin", color="AAAAAA"),
        )

        EXPORT_COLUMNS = [
            ("ID",           36),
            ("Num",          8),
            ("Question",     60),
            ("A",            30),
            ("B",            30),
            ("C",            30),
            ("D",            30),
            ("E",            30),
            ("Correct",      12),
            ("Year",         8),
            ("Category",     20),
            ("Subcategory",  28),
            ("Tag",          16),
            ("Source",       14),
            ("Similarity",   12),
            ("Ref #",        10),
            ("Ref Question", 60),
            ("Ref Correct",  14),
            ("Ref Year",     10),
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Custom Export"

        # Header row
        for ci, (name, width) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font      = font_header
            cell.fill      = fill_header
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Data rows
        for ri, rec in enumerate(records, start=2):
            src  = rec.get("source_qcm_full") or rec.get("source_qcm", {})
            bm   = rec["best_match"]
            ref  = bm.get("ref_qcm", {})
            score = bm["similarity"]

            # Tag: normalize to string
            tag_val = src.get("Tag", "")
            if isinstance(tag_val, list):
                tag_val = ", ".join(str(t) for t in tag_val)

            # Color coding
            if score == 0.0:
                row_fill = fill_grey
                sim_label = "No match"
            elif score >= self.color_green:
                row_fill = fill_green
                sim_label = f"{score:.1%}"
            elif score >= self.color_yellow:
                row_fill = fill_yellow
                sim_label = f"{score:.1%}"
            else:
                row_fill = fill_red
                sim_label = f"{score:.1%}"

            row_values = [
                rec.get("qcm_id", ""),
                src.get("Num", ""),
                src.get("Text", ""),
                src.get("A", ""),
                src.get("B", ""),
                src.get("C", ""),
                src.get("D", ""),
                src.get("E", ""),
                src.get("Correct", ""),
                src.get("Year", ""),
                src.get("categoryName", ""),
                src.get("subcategoryName", ""),
                tag_val,
                src.get("Source", ""),
                sim_label,
                ref.get("Num", "") if ref else "",
                ref.get("Text", "") if ref else "",
                ref.get("Correct", "") if ref else "",
                ref.get("Year", "") if ref else "",
            ]

            wrap_cols = {3, 17}  # Question, Ref Question
            for ci, val in enumerate(row_values, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill      = row_fill
                cell.font      = font_regular
                cell.border    = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(ci in wrap_cols)
                )

        # Legend sheet
        ls = wb.create_sheet("Legend")
        ls.column_dimensions["A"].width = 20
        ls.column_dimensions["B"].width = 55
        for row_data in [
            ("Color",      "Meaning"),
            ("🟢 Green",   f"Similarity ≥ {self.color_green:.0%}  (high match)"),
            ("🟡 Yellow",  f"Similarity {self.color_yellow:.0%}–{self.color_green:.0%}  (moderate match)"),
            ("🔴 Red",     f"Similarity < {self.color_yellow:.0%}  (weak match)"),
            ("⬜ Grey",    "No match found above threshold"),
        ]:
            ls.append(list(row_data))

        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}1"

        try:
            wb.save(path)
            print(f"\n✅ Custom XLSX saved → {path.name}  ({len(records)} QCMs)")
        except Exception as e:
            print(f"⚠️  Could not save custom XLSX: {e}")

    def export_from_existing(self) -> bool:
        """Re-trigger custom export from an already-saved step8_matches.json (no re-matching)."""
        output_dir = self._get_output_dir()
        json_path  = output_dir / "step8_matches.json"

        if not json_path.exists():
            print("❌ No step8_matches.json found. Please run Step 8 first.")
            return False

        with open(json_path, "r", encoding="utf-8") as f:
            all_records: List[Dict] = json.load(f)

        print(f"\n📂 Loaded {len(all_records)} records from {json_path.name}")
        self._offer_custom_export(all_records, output_dir)
        return True
