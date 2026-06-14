"""
xlsx_exporter.py
Shared utility to export a list of QCM dicts to a formatted .xlsx file
that matches the Template.xlsx column layout.

Usage:
    from modules.utils.xlsx_exporter import export_qcms_to_xlsx
    export_qcms_to_xlsx(qcm_list, Path("output/.../merged_qcms.xlsx"))
"""

import json
from pathlib import Path
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Style constants ──────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79") if HAS_OPENPYXL else None
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11) if HAS_OPENPYXL else None
ROW_FILL_ODD  = PatternFill("solid", fgColor="DCE6F1") if HAS_OPENPYXL else None
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF") if HAS_OPENPYXL else None
BORDER_THIN   = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
) if HAS_OPENPYXL else None

# ── Preferred column order ──────────────────────────────────────────────────
# Cas is included here so it appears right after Num when present.
# Any extra fields found in the QCM data are appended dynamically at the end.
_PREFERRED_COLUMNS = [
    "Num", "Cas", "Text", "A", "B", "C", "D", "E",
    "Correct", "Exp",
    "categoryName", "tagSuggere", "subcategoryName",
    "Year", "Tag", "Type"
]

# Approximate column widths (characters)
COL_WIDTHS = {
    "Num": 6, "Cas": 60, "Text": 60,
    "A": 40, "B": 40, "C": 40, "D": 40, "E": 40,
    "Correct": 10, "Exp": 40,
    "categoryName": 25, "tagSuggere": 15, "subcategoryName": 35,
    "Year": 8, "Tag": 20, "Type": 8,
}


def _build_columns(qcms: List[Dict]) -> List[str]:
    """Build the final column list dynamically.
    
    - Keeps _PREFERRED_COLUMNS order for known fields
    - Only includes a preferred column if at least one QCM has it (non-null)
    - Appends any extra fields found in the data that are not in the preferred list
    """
    # Gather all keys present in any QCM
    all_keys: set = set()
    for qcm in qcms:
        all_keys.update(k for k, v in qcm.items() if v is not None and v != "")
    
    # Preferred columns first (only those that actually appear in data)
    result = [c for c in _PREFERRED_COLUMNS if c in all_keys]
    
    # Append any extra keys not already in the preferred list
    for k in all_keys:
        if k not in result:
            result.append(k)
    
    return result



def _format_tag(value: Any) -> str:
    """Normalise Tag field: list -> "['2024']" string (matches Template.json format)."""
    if isinstance(value, list):
        return str(value)
    if value is None:
        return ""
    return str(value)


def _cell_value(qcm: Dict, col: str) -> Any:
    """Extract the right value for a column, with fallbacks."""
    val = qcm.get(col)
    if col == "Tag":
        return _format_tag(val)
    if val is None:
        return ""
    # Flatten lists to a comma-separated string (e.g., categoryName stored as list)
    if isinstance(val, list):
        # Deduplicate while preserving order, then join
        seen = []
        for item in val:
            s = str(item)
            if s not in seen:
                seen.append(s)
        return ", ".join(seen)
    return val


def export_qcms_to_xlsx(qcms: List[Dict], output_path: Path) -> bool:
    """
    Export a list of QCM dicts to a formatted .xlsx file.

    Args:
        qcms:        List of QCM dicts (from step5/6/7 JSON output).
        output_path: Where to save the .xlsx file (Path object).

    Returns:
        True on success, False on failure (e.g. openpyxl not installed).
    """
    if not HAS_OPENPYXL:
        print("⚠️  openpyxl not installed — skipping XLSX export. Run: pip install openpyxl")
        return False

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "QCMs"

    # ── Build dynamic column list ─────────────────────────────────────────────
    columns = _build_columns(qcms)

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER_THIN
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # Freeze header

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, qcm in enumerate(qcms, start=2):
        fill = ROW_FILL_ODD if row_idx % 2 == 0 else ROW_FILL_EVEN
        for col_idx, col_name in enumerate(columns, start=1):
            val = _cell_value(qcm, col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = BORDER_THIN
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=(col_name in ("Text", "Cas", "A", "B", "C", "D", "E", "Exp"))
            )

    # ── Auto-filter ──────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    try:
        wb.save(output_path)
        print(f"[XLSX] Saved -> {output_path.name}")
        return True
    except Exception as e:
        print(f"[WARN] Could not save XLSX: {e}")
        return False
